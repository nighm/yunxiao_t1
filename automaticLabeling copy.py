#!/usr/bin/python
# -*- coding: UTF-8 -*-

import sys
import os
import shutil
from datetime import datetime
import time
import yxConfig
import openpyxl
from openpyxl import Workbook
import re

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import random
from selenium.webdriver import ActionChains
from selenium.webdriver import ChromeOptions
# 设置浏览器,防止selenium被检测出来
options = ChromeOptions()
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_experimental_option('excludeSwitches', ['enable-automation'])

# 最后修改时间2023/12/14  Tag: 优化： 增加判断测试单中有不同产品用例库
# 最后修改时间2024/03/18  Tag: 优化： 增加挑选用例时的搜索模式--云效编号或用例标题（原因：云效更新）
# 最后修改时间2024/03/26  Tag: 增加功能： 增加测试用例引用模式
# 最后修改时间2024/05/11  Tag: 优化功能： 适配，云效更新导致的挑选用例模式选择用例库后没用例显示的问题
# 最后修改时间2024/05/22  Tag: 优化功能： 挑选用例模式测试单用例为空时增加判断
# 最后修改时间2024/07/09  Tag: 修复Bug： 挑选用例模式，云效编号筛选出多个用例时，可精准选择特定用例
# 最后修改时间2024/10/21  Tag: 修复Bug： 优化用例导出模式，适配超多失败用例及暂缓用例的情况



# 作者：娄武剑
# 测试结果未标记 列表
global list2
list2 = []
# 自动化类型 未标注用例列表
global noLabellist
noLabellist = []
# 标注模式
autoLabel = yxConfig.autoLabel
print(autoLabel)
# 产品名称
yxProductName = yxConfig.yxProductName
# 旧的测试单名称
autoPlanName = yxConfig.autoPlanName
# 新的测试单名称
newPlanName = yxConfig.newPlanName
# 挑选用例文件名
selectCaseFileName = yxConfig.selectCaseFileName
# 行号
lineNum = yxConfig.lineNum


# class automaticLabeling(WebAw.WebAw):
class automaticLabeling():

    def __init__(self):
        super().__init__()
        self.name = yxConfig.userName
        self.pwd = yxConfig.password
        # 打开谷歌浏览器
        # self.driver = webdriver.Chrome(
        #     executable_path='./Chrome/chromedriver.exe', options=options)
        self.driver = webdriver.Chrome(
            executable_path='./chromedriver/chromedriver-win64/chromedriver.exe')
        # 设置隐式等待时间
        self.driver.implicitly_wait(1.5)

    '''
       函数名：login
       说明：登录云效 函数
       参数：
       返回：
       作者：louwujian
    '''

    def login(self, name=yxConfig.userName, pwd=yxConfig.password):
        #云效地址
        url = 'https://devops.aliyun.com/workbench?orgId=63e607799dee9309492bc382'
        print(f'[DEBUG] 开始登录流程，目标URL: {url}')

        # 防检测自动化
        self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
            Object.defineProperty(navigator, 'webdriver', {
              get: () => undefined
            })
          """
        })
        print('[DEBUG] 已设置防检测自动化')
        
        #最大化窗口
        self.driver.maximize_window()
        print('[DEBUG] 窗口已最大化')
        
        #打开网页
        self.driver.get(url)
        print('[DEBUG] 已打开目标网页')
        time.sleep(2)
        
        #进入登录frame
        try:
            self.driver.switch_to.frame("alibaba-login-box")
            print('[DEBUG] 已切换到登录frame')
        except Exception as e:
            print(f'[ERROR] 切换到登录frame失败: {str(e)}')
            return False
            
        time.sleep(1)
        
        #输入用户名
        try:
            username_input = self.driver.find_element_by_xpath('//*[@id="fm-login-id"]')
            username_input.send_keys(name)
            print('[DEBUG] 已输入用户名')
        except Exception as e:
            print(f'[ERROR] 输入用户名失败: {str(e)}')
            return False
            
        time.sleep(1)
        
        #输入密码
        try:
            password_input = self.driver.find_element_by_xpath('//*[@id="fm-login-password"]')
            password_input.send_keys(pwd)
            print('[DEBUG] 已输入密码')
        except Exception as e:
            print(f'[ERROR] 输入密码失败: {str(e)}')
            return False
            
        time.sleep(1)
        
        #点击登录
        try:
            login_button = self.driver.find_element_by_xpath('//*[@id="login-form"]/div[5]/button')
            login_button.click()
            print('[DEBUG] 已点击登录按钮(路径1)')
        except:
            try:
                login_button = self.driver.find_element_by_xpath('//*[@id="login-form"]/div[6]/button')
                login_button.click()
                print('[DEBUG] 已点击登录按钮(路径2)')
            except Exception as e:
                print(f'[ERROR] 点击登录按钮失败: {str(e)}')
                return False
                
        time.sleep(2)
        
        # 查看是否有滑块
        try:
            slidXpath = self.driver.find_element_by_id('baxia-dialog-content')
            print('[DEBUG] 检测到滑块验证')
            slid_exist = True
        except:
            slid_exist = False
            print('[DEBUG] 未检测到滑块验证')

        # 判断是否滑动滑块
        if slid_exist == True:
            print('[DEBUG] 开始处理滑块验证')
            # 进入滑块frame
            try:
                self.driver.switch_to.frame("baxia-dialog-content")
                print('[DEBUG] 已切换到滑块frame')
            except Exception as e:
                print(f'[ERROR] 切换到滑块frame失败: {str(e)}')
                return False
                
            # 获取滑块
            try:
                button = self.driver.find_element_by_xpath('//*[@aria-label="滑块"]')
                print('[DEBUG] 已找到滑块元素')
            except Exception as e:
                print(f'[ERROR] 查找滑块元素失败: {str(e)}')
                return False
                
            # 建立动作链
            action = ActionChains(self.driver)
            action.click_and_hold(button)
            count = random.randint(110, 120)
            print(f'[DEBUG] 滑块拖动参数: 每次移动{count}像素')
            
            try:
                for i in range(5):
                    action.move_by_offset(count, 0).perform()
                    time.sleep(0.3)
                action.release()
                print('[DEBUG] 滑块拖动完成')
                time.sleep(2)
            except Exception as e:
                print(f'[ERROR] 滑块拖动失败: {str(e)}')
                return False

        # 等待进入工作台
        print('[DEBUG] 开始等待进入工作台')
        for pagei in range(15):
            current_url = self.driver.current_url
            print(f'[DEBUG] 当前URL: {current_url}')
            
            homePageXpath = '/html/body/div[1]/header/div/div/div[1]/div[2]/a/div[2]'
            print(f'[DEBUG] 尝试查找工作台元素，XPath: {homePageXpath}')

            if pagei < 14:
                try:
                    slidXpath = self.driver.find_element_by_xpath(homePageXpath)
                    print('[DEBUG] 成功找到工作台元素')
                    print('[DEBUG] 进入云效工作台成功')
                    return True
                except Exception as e:
                    print(f'[DEBUG] 等待进入云效工作台{14-pagei}... 错误信息: {str(e)}')
                time.sleep(1)
            if pagei == 14:
                time.sleep(1)
                try:
                    slidXpath = self.driver.find_element_by_xpath(homePageXpath)
                    print('[DEBUG] 成功找到工作台元素')
                    print('[DEBUG] 进入云效工作台成功')
                    return True
                except Exception as e:
                    print(f'[ERROR] 进入云效工作台失败: {str(e)}')
                    return False
        # time.sleep(15)

    '''
       函数名：readCsv
       说明：读取csv文件 函数
       参数：
       返回：
       作者：louwujian
    '''
    def readCsv(self, fileName = None):
        if fileName is not None:
            f = open(fileName, "r")  # 设置文件对象
            data = f.readlines()  # 直接将文件中按行读到list里，效果与方法2一样
            print('读取 %s 文件'  % fileName)
            f.close()  # 关闭文件
            return data
        else:
            print("请输入文件名")

    '''
       函数名：getElementExistanceByxpath
       说明：判断用例是否筛选到 用例
       参数：
       返回：
       作者：louwujian
    '''

    # 判断是否未查询到用例
    def getElementExistanceByxpath(self):
        """通过元素id判断是否存在该元素"""
        global element_existance
        element_existance = True

        try:
            # 尝试寻找元素，如若没有找到则会抛出异常
            # print("找用例")
            element = self.driver.find_element_by_xpath('//*[text()="暂无内容"]')
            # print(element)
            print('未找到用例')
        except:
            element_existance = False
            print('有用例')

        return element_existance

    '''
       函数名：getExistanceByxpath
       说明：判断用例是否筛选到 元素
       参数：
       返回：
       作者：louwujian
    '''

    # 判断是否未查询到用例
    def getExistanceByxpath(self, xpath=None):
        """通过元素id判断是否存在该元素"""
        global element_exist
        element_exist = True

        try:
            # 尝试寻找元素，如若没有找到则会抛出异常
            # print("找用例")
            elementXpath = self.driver.find_element_by_xpath(xpath)
            # print(elementXpath)
            print('有元素')
        except:
            element_exist = False
            print('无元素')

        return element_exist

    '''
       函数名：resultAuto
       说明：自动化结果标记 函数
       参数：CaseId为测试用例编号，CaseResult为测试用例执行结果（PASS、FAIL、暂缓），Other为除结果外是否需要标记其他字段，
            TestUser为执行人名称，例如TestUser='娄武剑'
       返回：
       作者：louwujian
    '''
    def resultAuto(self, CaseId=None, CaseResult=None, Other=False, TestUser='娄武剑'):
        # 点击 进行筛选
        try:
            # 尝试寻找元素，如若没有找到则会抛出异常
            self.driver.find_element_by_xpath('/html/body/div[2]/main/header/section/section/section/span[2]/button').click()
            # print('第一种')
        except:
            self.driver.find_element_by_xpath('/html/body/div[3]/main/header/section/section/section/span[2]/button').click()
            # print('第二种')

        # time.sleep(1)
        WebDriverWait(self.driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '//*[text()="过滤"]')))  # 等待显示 过滤框

        # 在 测试用例编号输入内容
        element = self.driver.find_element_by_xpath(
            '//*[text()="测试用例编号"]/./../../span/input')
        element.send_keys(Keys.CONTROL, "a")  # 相当于ctrl + a快捷键全选-+*/8の
        time.sleep(0.5)
        element.send_keys(Keys.DELETE)  # 快捷键删除
        time.sleep(0.5)
        element.send_keys(CaseId)
        time.sleep(0.5)
        # 点击 过滤
        self.driver.find_element_by_xpath('//*[text()="过滤"]/./..').click()
        # print('点击 过滤')
        # time.sleep(2)
        WebDriverWait(self.driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div')))  # 等待显示 过滤结果
        Test.getElementExistanceByxpath()
        # print(element_existance)
        # 是否有用例 有
        if element_existance == False:
            # 有用例
            element1 = self.driver.find_element_by_xpath(
                '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody/tr/td[7]').text
            print('获取到的目前标记结果')
            print(element1)
            if element1 == '待测试':
                #     # # #点击 选择自动化类型的下拉菜单
                name = self.driver.find_element_by_xpath(
                    '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody/tr/td[7]/div/button').click()
                print(name)
                # time.sleep(1)
                WebDriverWait(self.driver, 15).until(
                    EC.element_to_be_clickable((By.XPATH,
                                                '//*[text()="已通过"]')))  # 等待显示 结果标记下拉框
                if CaseResult == 'PASS\n':
                    #     # #点击 选择自动化类型的下拉菜单的    是
                    self.driver.find_element_by_xpath('//*[text()="已通过"]').click()
                    time.sleep(0.5)
                    WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH,
                                                    '//*[contains(text(),"已通过")]')))  # 等待显示 结果标记加载
                    # 获取当前标记结果
                    element1 = self.driver.find_element_by_xpath(
                        '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody/tr/td[7]').text
                    print('获取到的目前标记结果')
                    print(element1)
                    if element1 == '已通过':
                        print('标记成功')
                    else:
                        print('标记失败')
                        list2.append(CaseId + ',标记已通过失败')
                elif CaseResult == 'FAIL\n':
                    #     # #点击 选择自动化类型的下拉菜单的
                    self.driver.find_element_by_xpath('//*[text()="未通过"]').click()
                    time.sleep(0.5)
                    WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH,
                                                    '//*[contains(text(),"未通过")]')))  # 等待显示 结果标记加载
                    # 获取当前标记结果
                    element1 = self.driver.find_element_by_xpath(
                        '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody/tr/td[7]').text
                    print('获取到的目前标记结果')
                    print(element1)
                    if element1 == '未通过':
                        print('标记成功')
                    else:
                        print('标记失败')
                        list2.append(CaseId + ',标记未通过失败')
                else:
                    #     # #点击 选择自动化类型的下拉菜单的
                    self.driver.find_element_by_xpath('//*[text()="暂缓"]').click()
                    time.sleep(0.5)
                    WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH,
                                                    '//*[contains(text(),"暂缓")]')))  # 等待显示 结果标记加载
                    # 获取当前标记结果
                    element1 = self.driver.find_element_by_xpath(
                        '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody/tr/td[7]').text
                    print('获取到的目前标记结果')
                    print(element1)
                    if element1 == '暂缓':
                        print('标记成功')
                    else:
                        print('标记失败')
                        list2.append(CaseId + ',标记暂缓失败')
                if Other == True:
                    # 点击 选择用例
                    # 点击 选择用例 2种xpath
                    try:
                        # 尝试寻找元素，如若没有找到则会抛出异常
                        self.driver.find_element_by_xpath(
                            '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody/tr').click()
                        # print('第一种用例xpath')
                    except:
                        self.driver.find_element_by_xpath(
                            '/html/body/div[2]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody/tr').click()
                    #     # print('点击 选择用例')
                    WebDriverWait(self.driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[text()="前置条件"]')))  # 等待进入用例描述
                    # 判断是否需要修改执行人
                    if TestUser is not None:
                        # 需修改执行人
                        operatorUser = self.driver.find_element_by_xpath(
                            '//*[@id="workitemAttachment"]/../div[2]/div[2]/div[2]/div/div/span/span[1]/span[1]/em').text
                        # 判断现在的执行人是不是需要的 如相等则不需要修改
                        if operatorUser != TestUser:
                            self.driver.find_element_by_xpath(
                                '//*[@id="workitemAttachment"]/../div[2]/div[2]/div[2]/div/div/span/span[1]/span[2]').click()
                            WebDriverWait(self.driver, 15).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@placeholder="请输入关键字"]')))  # 等待进入输入框
                            self.driver.find_element_by_xpath('//*[@placeholder="请输入关键字"]').send_keys(TestUser)
                            userListXpath = '//*[@class="uiless-member-mini-v2-members"]'
                            WebDriverWait(self.driver, 15).until(
                                EC.element_to_be_clickable((By.XPATH, userListXpath)))  # 等待 出现人员
                            time.sleep(1)
                            userList = self.driver.find_element_by_xpath(userListXpath).text
                            if TestUser in userList:
                                self.driver.find_element_by_xpath(
                                    '//*[@placeholder="请输入关键字"]/../../../div/div/div/div').click()
                            else:
                                # 未找到人员则不修改，收起人员列表
                                self.driver.find_element_by_xpath(
                                    '//*[@id="workitemAttachment"]/../div[2]/div[2]/div[2]/div/div/span/span[1]/span[2]').click()
                    # 点击 收起用例内容
                    self.driver.find_element_by_xpath(
                        '//*[@id="drawer-sidebar-workitemDetail"]/../div').click()
                    time.sleep(1)
            # 是否有用例 无
            else:
                print('加入******不标记列表，已有标记')
                list2.append(CaseId + ',已有标记')
                # print(list2)
                # 点击 收起用例内容
                # self.driver.find_element_by_xpath('//*[@id="drawer-sidebar-workitemDetail"]/../div').click()
                time.sleep(1)

        elif element_existance == True:
            print('加入未标记列表******')
            list2.append(CaseId + ',未找到用例')
            print(list2)

    '''
       函数名：labelAuto
       说明：自动化测试类型标记 函数
       参数：
       返回：
       作者：louwujian
    '''
    def labelAuto(self, CaseId=None, CaseType=None):
    # def labelAuto(self, CaseId=None, CaseType=None):
        # 点击 进行筛选
        try:
            # 尝试寻找元素，如若没有找到则会抛出异常
            self.driver.find_element_by_xpath(
                '/html/body/div[2]/main/header/section/section/section/span[2]/button').click()
        except:
            self.driver.find_element_by_xpath(
                '/html/body/div[3]/main/header/section/section/section/span[2]/button').click()
        time.sleep(1)
        # 在 测试用例编号输入内容

        element = self.driver.find_element_by_xpath(
            '//*[text()="测试用例编号"]/./../../span/input')
        # element = self.driver.find_element_by_xpath(
        #     '//*[@id="container"]/main/section/section[2]/section/div/div[1]/div[1]/div/div/div[8]/span/input')
        element.send_keys(Keys.CONTROL, "a")  # 相当于ctrl + a快捷键全选
        time.sleep(1)
        element.send_keys(Keys.DELETE)  # 快捷键删除
        time.sleep(1)
        element.send_keys(CaseId)
        time.sleep(1)
        # 点击 过滤
        self.driver.find_element_by_xpath(
            '//*[@id="container"]/main/section/section[2]/section/div/div[1]/div[2]/button').click()
        # print('点击 过滤')
        time.sleep(2)
        Test.getElementExistanceByxpath()
        # print(element_existance)
        if element_existance == False:
            # 点击 选择用例
            self.driver.find_element_by_xpath(
                '/html/body/div[2]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div[2]/div[2]/div[2]/table/tbody/tr').click()
            time.sleep(1.5)
            print('点击 选择用例')
            print(CaseType)
            # #点击 选择自动化类型的下拉菜单
            name = self.driver.find_element_by_xpath('//*[@id="workitemAttachment"]/../div[1]/div/div[8]/div[2]').click()
            print(name)
            time.sleep(1)
            if CaseType == '是\n':
                # 点击 选择自动化类型的下拉菜单的    是
                self.driver.find_element_by_xpath('//span[text()="是"]/./..').click()
                time.sleep(1.5)
                print('点击 选择自动化类型    是')
            elif CaseType == '否\n':
                # 点击 选择自动化类型的下拉菜单的    否
                self.driver.find_element_by_xpath('//span[text()="否"]/./..').click()
                time.sleep(1.5)
                print('点击 选择自动化类型    否')
            else:
                print('自动化类型错误！！！')
                print('加入未标记列表******')
                noLabellist.append(CaseId + ',类型错误')


            # 点击 收起用例内容
            self.driver.find_element_by_xpath('//*[@id="drawer-sidebar-workitemDetail"]/../div').click()
            time.sleep(0.5)
            # print('点击 收起用例内容')
            # driver.switch_to.default_content() # 切出
        elif element_existance == True:
            print('未找到用例******')
            print('加入未标记列表******')
            noLabellist.append(CaseId + ',未找到用例')
            time.sleep(0.5)
            print(noLabellist)
    '''
       函数名：getHeaderAttr
       说明：获取摸个元素的特性属性值
       参数：
       返回：
       作者：louwujian
    '''
    def getHeaderAttr(self, xpath=None, attr=None):
        if xpath is not None:
            headerName = self.driver.find_element_by_xpath(xpath)
            if attr is not None:
                headerNameAttr = headerName.get_attribute(attr)
                return headerNameAttr
            else:
                print('要获取的属性没有填写')
        else:
            print('没有填写Xpath')

    '''
       函数名：getModelCase
       说明：获取某个模块用例
       参数：
       返回：
       作者：louwujian
    '''
    def getModelCase(self, lineXLS, sheet, modelNameNum):
        trCaseList = []
        trCaseNum = 0
        itemScro = 0
        while trCaseNum < int(modelNameNum):
            # 获取用例框
            print('开始获取显示的用例数')
            try:
                tbodyXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody'
                WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, tbodyXpath)))
                tbody = self.driver.find_element_by_xpath(tbodyXpath)
            except:
                tbodyXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/div/div/table/tbody'
                WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, tbodyXpath)))
                tbody = self.driver.find_element_by_xpath(tbodyXpath)
            trList = tbody.find_elements_by_xpath('tr')
            trNum = len(trList)
            print('开始判断用例块模式')
            # 开始判断，用例块模式 (模块中用例总数大于 xpth获取到的数量为小用例块模式，否则为大用例块模式；一般大于50个用例为大用例块)
            if trNum == int(modelNameNum):
                print('小用例块模式：')
                print('***********用例模块中显示的用例数量:%s个**********' % trNum)
                for iTr in range(1, trNum + 1):
                    trXpath = tbodyXpath + f'/tr[{iTr}]'
                    tr = self.driver.find_element_by_xpath(trXpath)
                    trContent = tr.text.split('\n')
                    TrCase = tr.text.replace('\n', ' ').replace('\r', ' ')  # 替换换行符号
                    # print(f'用例名称：{TrCase}')
                    print(f'用例变形前：{trContent}')
                    # print(f'用例列表：{trCaseList}')
                    bugTag = trContent[4]
                    if TrCase not in trCaseList:
                        trCaseNum += 1
                        lineXLS += 1
                        trCaseList.append(TrCase)
                        if '缺陷' not in trContent[4]:
                            trContent.insert(4, '--')
                        else:
                            trContent[4] = '是'
                        trCaseContent = trContent[0:2] + trContent[5:8] + trContent[2:5]
                        # print(f'变形后列表：{trCaseContent}')
                        if trCaseContent[6] == '未通过' or trCaseContent[6] == '暂缓':
                            if '缺陷' in bugTag:
                                tdBugXpath = trXpath + '/td[6]/div'
                                # ---*****增加 将元素移到可视区域*****---
                                tdBug = self.driver.find_element_by_xpath(tdBugXpath)
                                self.driver.execute_script("arguments[0].scrollIntoView();", tdBug)
                                tdBug.click()

                                # self.driver.find_element_by_xpath(tdBugXpath).click()
                                WebDriverWait(self.driver, 15).until(
                                    EC.element_to_be_clickable((By.XPATH, '// *[text() = "缺陷标题"]')))  # 等待缺陷标题出现
                                try:
                                    bugBodyXpath = '//*[text()="缺陷标题"]/../../../../../../../div[2]/table/tbody'
                                    bugBody = self.driver.find_element_by_xpath(bugBodyXpath)
                                    bugNum = bugBody.find_elements_by_xpath('tr')
                                except:
                                    bugBodyXpath = '/html/body/div[7]/div/div/div[2]/div[2]/table/tbody'
                                    bugBody = self.driver.find_element_by_xpath(bugBodyXpath)
                                    bugNum = bugBody.find_elements_by_xpath('tr')
                                BugId = ''
                                for bugN in range(1, len(bugNum)+1):
                                    tdBug.click()

                                    # self.driver.find_element_by_xpath(tdBugXpath).click()
                                    WebDriverWait(self.driver, 15).until(
                                        EC.element_to_be_clickable((By.XPATH, '// *[text() = "缺陷标题"]')))  # 等待缺陷标题出现
                                    bugNXpath = bugBodyXpath + f'/tr[{bugN}]/td[1]'
                                    bugN = self.driver.find_element_by_xpath(bugNXpath)
                                    # ---*****增加 将元素移到可视区域*****---
                                    self.driver.execute_script("arguments[0].scrollIntoView();", bugN)
                                    bugN.click()
                                    time.sleep(1)
                                    # 切换进入新窗口
                                    self.driver.switch_to.window(self.driver.window_handles[-1])
                                    WebDriverWait(self.driver, 15).until(
                                        EC.element_to_be_clickable((By.XPATH, '//*[contains(text(),"首页")]')))  # 等待页面加载
                                    print('判断权限')
                                    Test.getExistanceByxpath(xpath='//*[text()="暂无权限"]')
                                    # 判断是否有权限
                                    if element_exist == True:
                                        print('暂无权限')
                                        BugId = '暂无权限'
                                        time.sleep(1)
                                        self.driver.close()
                                        # 切换回原来窗口
                                        self.driver.switch_to.window(self.driver.window_handles[0])
                                    else:
                                        print('有权限')
                                        print(f'当前页面title为：{self.driver.title}')
                                        WebDriverWait(self.driver, 15).until(
                                            EC.element_to_be_clickable((By.XPATH, '//*[text()="问题单号"]')))  # 等待问题单号出现
                                        BugValue = self.driver.find_element_by_xpath(
                                            '//*[text()="问题单号"]/../../div[2]/div/div/span/input').get_attribute('value')
                                        print(f'问题单号为：{BugValue}')
                                        BugId = BugId + BugValue + '\n'
                                        time.sleep(1)
                                        self.driver.close()
                                        # 切换回原来窗口
                                        self.driver.switch_to.window(self.driver.window_handles[0])
                                trCaseContent.append(BugId)
                            else:
                                trCaseContent.append('--')
                            tdXpath = trXpath + '/td[2]/div'
                            # ---*****增加 将元素移到可视区域*****---
                            tdBT = self.driver.find_element_by_xpath(tdXpath)
                            self.driver.execute_script("arguments[0].scrollIntoView();", tdBT)
                            tdBT.click()

                            # self.driver.find_element_by_xpath(tdXpath).click()
                            # print('进入用例描述')
                            try:
                                WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[text()="前置条件"]'))) # 等待进入用例描述
                            except:
                                # 刷新
                                self.driver.refresh()
                                WebDriverWait(self.driver, 15).until(
                                    EC.element_to_be_clickable((By.XPATH, '//*[text()="前置条件"]')))  # 等待进入用例描述
                            dynamicXpath = '//*[@id="workitemModules"]/div/div[2]/div/div/ul/li/div/div[3]/div[2]/div/div[1]/div/div'
                            Test.getExistanceByxpath(xpath=dynamicXpath)
                            # 判断是否有动态内容
                            if element_exist == True:
                                dynamic = self.driver.find_element_by_xpath(
                                    '//*[@id="workitemModules"]/div/div[2]/div/div/ul/li/div/div[3]/div[2]/div/div[1]/div/div').text
                                trCaseContent.append(dynamic)
                            else:
                                trCaseContent.append('--')
                            # 点击 收起用例内容
                            self.driver.find_element_by_xpath('//*[@id="drawer-sidebar-workitemDetail"]/../div').click()
                            time.sleep(1)
                            # tr.click()
                            for iLine in range(1,len(trCaseContent)+1):
                                sheet.cell(lineXLS, iLine, str(trCaseContent[iLine-1]))  # 将指定值写入第i+1行第j+1列
                            # print(f'用例为{trCaseContent}')
                            print(f'用例数量已增加为：{trCaseNum}')
                            print(f'excel行数已增加为：{lineXLS}')
                            #  如果显示用例小于模块用例遇到未通过和暂缓，break（因为xpath会动态变化）
                            if trNum < int(modelNameNum):
                                if trCaseNum < int(modelNameNum):
                                    break
                        else:
                            # tr.click()
                            for iLine in range(1,len(trCaseContent)+1):
                                sheet.cell(lineXLS, iLine, str(trCaseContent[iLine-1]))  # 将指定值写入第i+1行第j+1列
                            # print(f'用例为{trCaseContent}')
                            print(f'当前模块用例数量已增加为：{trCaseNum}')
                            print(f'excel行数已增加为：{lineXLS}')
                            #  如果显示用例小于模块用例遇到最后一个需要滚动滑块
                            if trNum < int(modelNameNum):
                                if trCaseNum < int(modelNameNum):
                                    if iTr == trNum:
                                        scrollbarXpath = tbodyXpath + f'/tr[{iTr}]' + '/td[1]/div/label/span[1]/input'
                                        # print(scrollbarXpath)
                                        scrollbar = self.driver.find_element_by_xpath(scrollbarXpath)
                                        scrollbar.click()
                                        scrollbar.send_keys(Keys.PAGE_DOWN)
                    else:
                        print('用例已存在，跳过！！')
            else:
                print('大用例块模式：')
                print('失败用例----->')
                # 点击 进行筛选
                try:
                    # 尝试寻找元素，如若没有找到则会抛出异常
                    self.driver.find_element_by_xpath(
                        '/html/body/div[2]/main/header/section/section/section/span[2]/button').click()
                    # print('第一种')
                except:
                    self.driver.find_element_by_xpath(
                        '/html/body/div[3]/main/header/section/section/section/span[2]/button').click()
                    # print('第二种')
                time.sleep(1)
                # 添加状态
                element = self.driver.find_element_by_xpath(
                    '//*[@id="container"]/main/section/section[2]/section/div/div[1]/div[1]/div/div/div[5]/div[3]').click()
                time.sleep(1.5)
                # 点击 暂缓
                # self.driver.find_element_by_xpath(
                #     '//*[@id="aoneCommonSearchDropdown"]/div[1]/ul/li[2]/div/label/span/input').click()
                time.sleep(0.5)
                # 点击 未通过
                self.driver.find_element_by_xpath(
                    '//*[@id="aoneCommonSearchDropdown"]/div[1]/ul/li[4]/div/label/span/input').click()
                time.sleep(1)
                self.driver.find_element_by_xpath('//*[text()="确定"]/..').click()
                time.sleep(1)
                # 点击 过滤
                self.driver.find_element_by_xpath('//*[text()="过滤"]/./..').click()
                time.sleep(1)
                WebDriverWait(self.driver, 15).until(
                    EC.element_to_be_clickable((By.XPATH,
                                                '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div')))  # 等待显示 过滤结果


                # 判断是否有用例
                Test.getElementExistanceByxpath()
                # 如果有用例
                if element_existance == False:

                    errorNum = 0
                    # 全选 用例
                    self.driver.find_element_by_xpath(
                        '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[1]/div/div[1]/label/span/input').click()
                    # 获取有多少用例
                    selectErrorXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[1]/div/div[1]/label/span/input/../../../../div[2]'
                    selectError = self.driver.find_element_by_xpath(selectErrorXpath).text
                    print(selectError)

                    # 取消全选 用例
                    self.driver.find_element_by_xpath(
                        '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[1]/div/div[1]/label/span/input').click()
                    # 处理数据
                    try:
                        selectErrorList = selectError.split(' ')
                        selectErrorNum = int(selectErrorList[1])
                        print(f'筛选后的用例数：{selectErrorNum}')
                    except:
                        selectErrorNum = int(re.findall(r'\d+', selectError)[0])
                        print(f'筛选后的用例数：{selectErrorNum}')
                    if selectErrorNum > 50:
                        while errorNum < selectErrorNum:
                            # 点击 选择用例
                            # 获取用例框

                            try:
                                tbodyXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody'
                                WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable((By.XPATH, tbodyXpath)))
                                tbody = self.driver.find_element_by_xpath(tbodyXpath)
                            except:
                                tbodyXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/div/div/table/tbody'
                                WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable((By.XPATH, tbodyXpath)))
                                tbody = self.driver.find_element_by_xpath(tbodyXpath)
                            trList = tbody.find_elements_by_xpath('tr')
                            trNum = len(trList)
                            print('***********用例模块中显示的用例数量:%s个**********' % trNum)
                            for iTr in range(1, trNum + 1):
                                trXpath = tbodyXpath + f'/tr[{iTr}]'
                                tr = self.driver.find_element_by_xpath(trXpath)
                                trContent = tr.text.split('\n')
                                TrCase = tr.text.replace('\n', ' ').replace('\r', ' ')  # 替换换行符号
                                # print(f'用例名称：{TrCase}')
                                print(f'用例变形前：{trContent}')
                                # print(f'用例列表：{trCaseList}')
                                bugTag = trContent[4]
                                if TrCase not in trCaseList:
                                    errorNum +=1
                                    trCaseNum += 1
                                    lineXLS += 1
                                    trCaseList.append(TrCase)
                                    if '缺陷' not in trContent[4]:
                                        trContent.insert(4, '--')
                                    else:
                                        trContent[4] = '是'
                                    trCaseContent = trContent[0:2] + trContent[5:8] + trContent[2:5]
                                    # print(f'变形后列表：{trCaseContent}')
                                    if trCaseContent[6] == '未通过' or trCaseContent[6] == '暂缓':
                                        if '缺陷' in bugTag:
                                            tdBugXpath = trXpath + '/td[6]/div'
                                            # ---*****增加 将元素移到可视区域*****---
                                            tdBug = self.driver.find_element_by_xpath(tdBugXpath)
                                            self.driver.execute_script("arguments[0].scrollIntoView();", tdBug)
                                            tdBug.click()

                                            # self.driver.find_element_by_xpath(tdBugXpath).click()
                                            WebDriverWait(self.driver, 15).until(
                                                EC.element_to_be_clickable((By.XPATH, '// *[text() = "缺陷标题"]')))  # 等待缺陷标题出现
                                            try:
                                                bugBodyXpath = '//*[text()="缺陷标题"]/../../../../../../../div[2]/table/tbody'
                                                bugBody = self.driver.find_element_by_xpath(bugBodyXpath)
                                                bugNum = bugBody.find_elements_by_xpath('tr')
                                            except:
                                                bugBodyXpath = '/html/body/div[7]/div/div/div[2]/div[2]/table/tbody'
                                                bugBody = self.driver.find_element_by_xpath(bugBodyXpath)
                                                bugNum = bugBody.find_elements_by_xpath('tr')
                                            BugId = ''
                                            for bugN in range(1, len(bugNum)+1):
                                                tdBug.click()

                                                # self.driver.find_element_by_xpath(tdBugXpath).click()
                                                WebDriverWait(self.driver, 15).until(
                                                    EC.element_to_be_clickable((By.XPATH, '// *[text() = "缺陷标题"]')))  # 等待缺陷标题出现
                                                bugNXpath = bugBodyXpath + f'/tr[{bugN}]/td[1]'
                                                bugN = self.driver.find_element_by_xpath(bugNXpath)
                                                # ---*****增加 将元素移到可视区域*****---
                                                self.driver.execute_script("arguments[0].scrollIntoView();", bugN)
                                                bugN.click()
                                                time.sleep(1)
                                                # 切换进入新窗口
                                                self.driver.switch_to.window(self.driver.window_handles[-1])
                                                WebDriverWait(self.driver, 15).until(
                                                    EC.element_to_be_clickable(
                                                        (By.XPATH, '//*[contains(text(),"首页")]')))  # 等待页面加载
                                                print('判断权限')
                                                Test.getExistanceByxpath(xpath='//*[text()="暂无权限"]')
                                                # 判断是否有权限
                                                if element_exist == True:
                                                    print('暂无权限')
                                                    BugId = '暂无权限'
                                                    time.sleep(1)
                                                    self.driver.close()
                                                    # 切换回原来窗口
                                                    self.driver.switch_to.window(self.driver.window_handles[0])
                                                else:
                                                    print('有权限')
                                                    print(f'当前页面title为：{self.driver.title}')
                                                    WebDriverWait(self.driver, 15).until(
                                                        EC.element_to_be_clickable((By.XPATH, '//*[text()="问题单号"]')))  # 等待问题单号出现
                                                    BugValue = self.driver.find_element_by_xpath(
                                                        '//*[text()="问题单号"]/../../div[2]/div/div/span/input').get_attribute('value')
                                                    print(f'问题单号为：{BugValue}')
                                                    BugId = BugId + BugValue + '\n'
                                                    time.sleep(1)
                                                    self.driver.close()
                                                    # 切换回原来窗口
                                                    self.driver.switch_to.window(self.driver.window_handles[0])
                                            trCaseContent.append(BugId)
                                        else:
                                            trCaseContent.append('--')
                                        # 此处点击进入用例无需优化，因为已经可点击缺陷，此处应该可点击
                                        tdXpath = trXpath + '/td[2]/div'

                                        self.driver.find_element_by_xpath(tdXpath).click()
                                        # print('进入用例描述')
                                        try:
                                            WebDriverWait(self.driver, 15).until(
                                                EC.element_to_be_clickable((By.XPATH, '//*[text()="前置条件"]')))  # 等待进入用例描述
                                        except:
                                            # 刷新
                                            self.driver.refresh()
                                            WebDriverWait(self.driver, 15).until(
                                                EC.element_to_be_clickable((By.XPATH, '//*[text()="前置条件"]')))  # 等待进入用例描述
                                        dynamicXpath = '//*[@id="workitemModules"]/div/div[2]/div/div/ul/li/div/div[3]/div[2]/div/div[1]/div/div'
                                        Test.getExistanceByxpath(xpath=dynamicXpath)
                                        # 判断是否有动态内容
                                        if element_exist == True:
                                            dynamic = self.driver.find_element_by_xpath(
                                                '//*[@id="workitemModules"]/div/div[2]/div/div/ul/li/div/div[3]/div[2]/div/div[1]/div/div').text
                                            trCaseContent.append(dynamic)
                                        else:
                                            trCaseContent.append('--')
                                        # 点击 收起用例内容
                                        self.driver.find_element_by_xpath('//*[@id="drawer-sidebar-workitemDetail"]/../div').click()
                                        time.sleep(1)
                                        # tr.click()
                                        for iLine in range(1,len(trCaseContent)+1):
                                            sheet.cell(lineXLS, iLine, str(trCaseContent[iLine-1]))  # 将指定值写入第i+1行第j+1列
                                        # print(f'用例为{trCaseContent}')
                                        print(f'用例数量已增加为：{trCaseNum}')
                                        print(f'excel行数已增加为：{lineXLS}')


                                        if iTr >= 15:
                                        # 如果显示用例小于模块用例遇到未通过和暂缓，break（因为xpath会动态变化）
                                            if trNum < int(modelNameNum):
                                                if trCaseNum < int(modelNameNum):
                                                    break
                                    else:
                                        # tr.click()
                                        for iLine in range(1,len(trCaseContent)+1):
                                            sheet.cell(lineXLS, iLine, str(trCaseContent[iLine-1]))  # 将指定值写入第i+1行第j+1列
                                        # print(f'用例为{trCaseContent}')
                                        print(f'当前模块用例数量已增加为：{trCaseNum}')
                                        print(f'excel行数已增加为：{lineXLS}')
                                        #  如果显示用例小于模块用例遇到最后一个需要滚动滑块
                                        if trNum < int(modelNameNum):
                                            if trCaseNum < int(modelNameNum):
                                                if iTr == trNum:
                                                    scrollbarXpath = tbodyXpath + f'/tr[{iTr}]' + '/td[1]/div/label/span[1]/input'
                                                    # print(scrollbarXpath)
                                                    scrollbar = self.driver.find_element_by_xpath(scrollbarXpath)
                                                    scrollbar.click()
                                                    scrollbar.send_keys(Keys.PAGE_DOWN)
                                else:
                                    print('用例已存在，跳过！！')
                            time.sleep(1)
                    else:
                        # 点击 选择用例
                        # 获取用例框
                        try:
                            tbodyXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody'
                            WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable((By.XPATH, tbodyXpath)))
                            tbody = self.driver.find_element_by_xpath(tbodyXpath)
                        except:
                            tbodyXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/div/div/table/tbody'
                            WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable((By.XPATH, tbodyXpath)))
                            tbody = self.driver.find_element_by_xpath(tbodyXpath)
                        trList = tbody.find_elements_by_xpath('tr')
                        trNum = len(trList)
                        print('***********用例模块中显示的用例数量:%s个**********' % trNum)
                        for iTr in range(1, trNum + 1):
                            trXpath = tbodyXpath + f'/tr[{iTr}]'
                            tr = self.driver.find_element_by_xpath(trXpath)
                            trContent = tr.text.split('\n')
                            TrCase = tr.text.replace('\n', ' ').replace('\r', ' ')  # 替换换行符号
                            # print(f'用例名称：{TrCase}')
                            print(f'用例变形前：{trContent}')
                            # print(f'用例列表：{trCaseList}')
                            bugTag = trContent[4]
                            if TrCase not in trCaseList:
                                trCaseNum += 1
                                lineXLS += 1
                                trCaseList.append(TrCase)
                                if '缺陷' not in trContent[4]:
                                    trContent.insert(4, '--')
                                else:
                                    trContent[4] = '是'
                                trCaseContent = trContent[0:2] + trContent[5:8] + trContent[2:5]
                                # print(f'变形后列表：{trCaseContent}')
                                if trCaseContent[6] == '未通过' or trCaseContent[6] == '暂缓':
                                    if '缺陷' in bugTag:
                                        tdBugXpath = trXpath + '/td[6]/div'
                                        # ---*****增加 将元素移到可视区域*****---
                                        tdBug = self.driver.find_element_by_xpath(tdBugXpath)
                                        self.driver.execute_script("arguments[0].scrollIntoView();", tdBug)
                                        tdBug.click()

                                        # self.driver.find_element_by_xpath(tdBugXpath).click()
                                        WebDriverWait(self.driver, 15).until(
                                            EC.element_to_be_clickable(
                                                (By.XPATH, '// *[text() = "缺陷标题"]')))  # 等待缺陷标题出现
                                        try:
                                            bugBodyXpath = '//*[text()="缺陷标题"]/../../../../../../../div[2]/table/tbody'
                                            bugBody = self.driver.find_element_by_xpath(bugBodyXpath)
                                            bugNum = bugBody.find_elements_by_xpath('tr')
                                        except:
                                            bugBodyXpath = '/html/body/div[7]/div/div/div[2]/div[2]/table/tbody'
                                            bugBody = self.driver.find_element_by_xpath(bugBodyXpath)
                                            bugNum = bugBody.find_elements_by_xpath('tr')
                                        BugId = ''
                                        for bugN in range(1, len(bugNum) + 1):
                                            tdBug.click()

                                            # self.driver.find_element_by_xpath(tdBugXpath).click()
                                            WebDriverWait(self.driver, 15).until(
                                                EC.element_to_be_clickable(
                                                    (By.XPATH, '// *[text() = "缺陷标题"]')))  # 等待缺陷标题出现
                                            bugNXpath = bugBodyXpath + f'/tr[{bugN}]/td[1]'
                                            bugN = self.driver.find_element_by_xpath(bugNXpath)
                                            # ---*****增加 将元素移到可视区域*****---
                                            self.driver.execute_script("arguments[0].scrollIntoView();", bugN)
                                            bugN.click()
                                            time.sleep(1)
                                            # 切换进入新窗口
                                            self.driver.switch_to.window(self.driver.window_handles[-1])
                                            WebDriverWait(self.driver, 15).until(
                                                EC.element_to_be_clickable(
                                                    (By.XPATH, '//*[contains(text(),"首页")]')))  # 等待页面加载
                                            print('判断权限')
                                            Test.getExistanceByxpath(xpath='//*[text()="暂无权限"]')
                                            # 判断是否有权限
                                            if element_exist == True:
                                                print('暂无权限')
                                                BugId = '暂无权限'
                                                time.sleep(1)
                                                self.driver.close()
                                                # 切换回原来窗口
                                                self.driver.switch_to.window(self.driver.window_handles[0])
                                            else:
                                                print('有权限')
                                                print(f'当前页面title为：{self.driver.title}')
                                                WebDriverWait(self.driver, 15).until(
                                                    EC.element_to_be_clickable(
                                                        (By.XPATH, '//*[text()="问题单号"]')))  # 等待问题单号出现
                                                BugValue = self.driver.find_element_by_xpath(
                                                    '//*[text()="问题单号"]/../../div[2]/div/div/span/input').get_attribute(
                                                    'value')
                                                print(f'问题单号为：{BugValue}')
                                                BugId = BugId + BugValue + '\n'
                                                time.sleep(1)
                                                self.driver.close()
                                                # 切换回原来窗口
                                                self.driver.switch_to.window(self.driver.window_handles[0])
                                        trCaseContent.append(BugId)
                                    else:
                                        trCaseContent.append('--')
                                    # 此处点击进入用例无需优化，因为已经可点击缺陷，此处应该可点击
                                    tdXpath = trXpath + '/td[2]/div'

                                    self.driver.find_element_by_xpath(tdXpath).click()
                                    # print('进入用例描述')
                                    try:
                                        WebDriverWait(self.driver, 15).until(
                                            EC.element_to_be_clickable(
                                                (By.XPATH, '//*[text()="前置条件"]')))  # 等待进入用例描述
                                    except:
                                        # 刷新
                                        self.driver.refresh()
                                        WebDriverWait(self.driver, 15).until(
                                            EC.element_to_be_clickable(
                                                (By.XPATH, '//*[text()="前置条件"]')))  # 等待进入用例描述
                                    dynamicXpath = '//*[@id="workitemModules"]/div/div[2]/div/div/ul/li/div/div[3]/div[2]/div/div[1]/div/div'
                                    Test.getExistanceByxpath(xpath=dynamicXpath)
                                    # 判断是否有动态内容
                                    if element_exist == True:
                                        dynamic = self.driver.find_element_by_xpath(
                                            '//*[@id="workitemModules"]/div/div[2]/div/div/ul/li/div/div[3]/div[2]/div/div[1]/div/div').text
                                        trCaseContent.append(dynamic)
                                    else:
                                        trCaseContent.append('--')
                                    # 点击 收起用例内容
                                    self.driver.find_element_by_xpath(
                                        '//*[@id="drawer-sidebar-workitemDetail"]/../div').click()
                                    time.sleep(1)
                                    # tr.click()
                                    for iLine in range(1, len(trCaseContent) + 1):
                                        sheet.cell(lineXLS, iLine,
                                                   str(trCaseContent[iLine - 1]))  # 将指定值写入第i+1行第j+1列
                                    # print(f'用例为{trCaseContent}')
                                    print(f'用例数量已增加为：{trCaseNum}')
                                    print(f'excel行数已增加为：{lineXLS}')
                                    # if iTr >= 15:
                                    #  如果显示用例小于模块用例遇到未通过和暂缓，break（因为xpath会动态变化）
                                    # if trNum < int(modelNameNum):
                                    #     if trCaseNum < int(modelNameNum):
                                    #         break
                                else:
                                    # tr.click()
                                    for iLine in range(1, len(trCaseContent) + 1):
                                        sheet.cell(lineXLS, iLine,
                                                   str(trCaseContent[iLine - 1]))  # 将指定值写入第i+1行第j+1列
                                    # print(f'用例为{trCaseContent}')
                                    print(f'当前模块用例数量已增加为：{trCaseNum}')
                                    print(f'excel行数已增加为：{lineXLS}')
                                    #  如果显示用例小于模块用例遇到最后一个需要滚动滑块
                                    if trNum < int(modelNameNum):
                                        if trCaseNum < int(modelNameNum):
                                            if iTr == trNum:
                                                scrollbarXpath = tbodyXpath + f'/tr[{iTr}]' + '/td[1]/div/label/span[1]/input'
                                                # print(scrollbarXpath)
                                                scrollbar = self.driver.find_element_by_xpath(scrollbarXpath)
                                                scrollbar.click()
                                                scrollbar.send_keys(Keys.PAGE_DOWN)
                            else:
                                print('用例已存在，跳过！！')
                        time.sleep(1)
                    self.driver.find_element_by_xpath('//*[text()="清空"]/..').click()
                # 如果没有用例
                else:
                    time.sleep(1)
                    self.driver.find_element_by_xpath('//*[text()="清空"]/..').click()

                print('暂缓用例----->')
                # 点击 进行筛选
                try:
                    # 尝试寻找元素，如若没有找到则会抛出异常
                    self.driver.find_element_by_xpath(
                        '/html/body/div[2]/main/header/section/section/section/span[2]/button').click()
                    # print('第一种')
                except:
                    self.driver.find_element_by_xpath(
                        '/html/body/div[3]/main/header/section/section/section/span[2]/button').click()
                    # print('第二种')
                time.sleep(1)
                # 添加状态
                element = self.driver.find_element_by_xpath(
                    '//*[@id="container"]/main/section/section[2]/section/div/div[1]/div[1]/div/div/div[5]/div[3]').click()
                time.sleep(1.5)
                # 点击 暂缓
                self.driver.find_element_by_xpath(
                    '//*[@id="aoneCommonSearchDropdown"]/div[1]/ul/li[2]/div/label/span/input').click()
                time.sleep(0.5)
                # 点击 未通过
                # self.driver.find_element_by_xpath(
                #     '//*[@id="aoneCommonSearchDropdown"]/div[1]/ul/li[4]/div/label/span/input').click()
                time.sleep(1)
                self.driver.find_element_by_xpath('//*[text()="确定"]/..').click()
                time.sleep(1)
                # 点击 过滤
                self.driver.find_element_by_xpath('//*[text()="过滤"]/./..').click()
                time.sleep(1)
                WebDriverWait(self.driver, 15).until(
                    EC.element_to_be_clickable((By.XPATH,
                                                '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div')))  # 等待显示 过滤结果

                # 判断是否有用例
                Test.getElementExistanceByxpath()
                # 如果有用例
                if element_existance == False:

                    delayNum = 0
                    # 全选 用例
                    self.driver.find_element_by_xpath(
                        '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[1]/div/div[1]/label/span/input').click()
                    # 获取有多少用例
                    selectDelayXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[1]/div/div[1]/label/span/input/../../../../div[2]'
                    selectDelay = self.driver.find_element_by_xpath(selectDelayXpath).text
                    print(selectDelay)

                    # 取消全选 用例
                    self.driver.find_element_by_xpath(
                        '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[1]/div/div[1]/label/span/input').click()
                    # 处理数据
                    try:
                        selectDelayList = selectDelay.split(' ')
                        selectDelayNum = int(selectDelayList[1])
                        print(f'筛选后的用例数：{selectDelayNum}')
                    except:
                        selectDelayNum = int(re.findall(r'\d+', selectDelay)[0])
                        print(f'筛选后的用例数：{selectDelayNum}')
                    if selectDelayNum > 50:
                        while delayNum < selectDelayNum:
                            # 点击 选择用例
                            # 获取用例框
                            try:
                                tbodyXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody'
                                WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable((By.XPATH, tbodyXpath)))
                                tbody = self.driver.find_element_by_xpath(tbodyXpath)
                            except:
                                tbodyXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/div/div/table/tbody'
                                WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable((By.XPATH, tbodyXpath)))
                                tbody = self.driver.find_element_by_xpath(tbodyXpath)
                            trList = tbody.find_elements_by_xpath('tr')
                            trNum = len(trList)
                            print('***********用例模块中显示的用例数量:%s个**********' % trNum)
                            for iTr in range(1, trNum + 1):
                                trXpath = tbodyXpath + f'/tr[{iTr}]'
                                tr = self.driver.find_element_by_xpath(trXpath)
                                trContent = tr.text.split('\n')
                                TrCase = tr.text.replace('\n', ' ').replace('\r', ' ')  # 替换换行符号
                                # print(f'用例名称：{TrCase}')
                                print(f'用例变形前：{trContent}')
                                # print(f'用例列表：{trCaseList}')
                                bugTag = trContent[4]
                                if TrCase not in trCaseList:
                                    delayNum += 1
                                    trCaseNum += 1
                                    lineXLS += 1
                                    trCaseList.append(TrCase)
                                    if '缺陷' not in trContent[4]:
                                        trContent.insert(4, '--')
                                    else:
                                        trContent[4] = '是'
                                    trCaseContent = trContent[0:2] + trContent[5:8] + trContent[2:5]
                                    # print(f'变形后列表：{trCaseContent}')
                                    if trCaseContent[6] == '未通过' or trCaseContent[6] == '暂缓':
                                        if '缺陷' in bugTag:
                                            tdBugXpath = trXpath + '/td[6]/div'
                                            # ---*****增加 将元素移到可视区域*****---
                                            tdBug = self.driver.find_element_by_xpath(tdBugXpath)
                                            self.driver.execute_script("arguments[0].scrollIntoView();", tdBug)
                                            tdBug.click()

                                            # self.driver.find_element_by_xpath(tdBugXpath).click()
                                            WebDriverWait(self.driver, 15).until(
                                                EC.element_to_be_clickable(
                                                    (By.XPATH, '// *[text() = "缺陷标题"]')))  # 等待缺陷标题出现
                                            try:
                                                bugBodyXpath = '//*[text()="缺陷标题"]/../../../../../../../div[2]/table/tbody'
                                                bugBody = self.driver.find_element_by_xpath(bugBodyXpath)
                                                bugNum = bugBody.find_elements_by_xpath('tr')
                                            except:
                                                bugBodyXpath = '/html/body/div[7]/div/div/div[2]/div[2]/table/tbody'
                                                bugBody = self.driver.find_element_by_xpath(bugBodyXpath)
                                                bugNum = bugBody.find_elements_by_xpath('tr')
                                            BugId = ''
                                            for bugN in range(1, len(bugNum) + 1):
                                                tdBug.click()

                                                # self.driver.find_element_by_xpath(tdBugXpath).click()
                                                WebDriverWait(self.driver, 15).until(
                                                    EC.element_to_be_clickable(
                                                        (By.XPATH, '// *[text() = "缺陷标题"]')))  # 等待缺陷标题出现
                                                bugNXpath = bugBodyXpath + f'/tr[{bugN}]/td[1]'
                                                bugN = self.driver.find_element_by_xpath(bugNXpath)
                                                # ---*****增加 将元素移到可视区域*****---
                                                self.driver.execute_script("arguments[0].scrollIntoView();", bugN)
                                                bugN.click()
                                                time.sleep(1)
                                                # 切换进入新窗口
                                                self.driver.switch_to.window(self.driver.window_handles[-1])
                                                WebDriverWait(self.driver, 15).until(
                                                    EC.element_to_be_clickable(
                                                        (By.XPATH, '//*[contains(text(),"首页")]')))  # 等待页面加载
                                                print('判断权限')
                                                Test.getExistanceByxpath(xpath='//*[text()="暂无权限"]')
                                                # 判断是否有权限
                                                if element_exist == True:
                                                    print('暂无权限')
                                                    BugId = '暂无权限'
                                                    time.sleep(1)
                                                    self.driver.close()
                                                    # 切换回原来窗口
                                                    self.driver.switch_to.window(self.driver.window_handles[0])
                                                else:
                                                    print('有权限')
                                                    print(f'当前页面title为：{self.driver.title}')
                                                    WebDriverWait(self.driver, 15).until(
                                                        EC.element_to_be_clickable(
                                                            (By.XPATH, '//*[text()="问题单号"]')))  # 等待问题单号出现
                                                    BugValue = self.driver.find_element_by_xpath(
                                                        '//*[text()="问题单号"]/../../div[2]/div/div/span/input').get_attribute(
                                                        'value')
                                                    print(f'问题单号为：{BugValue}')
                                                    BugId = BugId + BugValue + '\n'
                                                    time.sleep(1)
                                                    self.driver.close()
                                                    # 切换回原来窗口
                                                    self.driver.switch_to.window(self.driver.window_handles[0])
                                            trCaseContent.append(BugId)
                                        else:
                                            trCaseContent.append('--')
                                        # 此处点击进入用例无需优化，因为已经可点击缺陷，此处应该可点击
                                        tdXpath = trXpath + '/td[2]/div'

                                        self.driver.find_element_by_xpath(tdXpath).click()
                                        # print('进入用例描述')
                                        try:
                                            WebDriverWait(self.driver, 15).until(
                                                EC.element_to_be_clickable((By.XPATH, '//*[text()="前置条件"]')))  # 等待进入用例描述
                                        except:
                                            # 刷新
                                            self.driver.refresh()
                                            WebDriverWait(self.driver, 15).until(
                                                EC.element_to_be_clickable((By.XPATH, '//*[text()="前置条件"]')))  # 等待进入用例描述
                                        dynamicXpath = '//*[@id="workitemModules"]/div/div[2]/div/div/ul/li/div/div[3]/div[2]/div/div[1]/div/div'
                                        Test.getExistanceByxpath(xpath=dynamicXpath)
                                        # 判断是否有动态内容
                                        if element_exist == True:
                                            dynamic = self.driver.find_element_by_xpath(
                                                '//*[@id="workitemModules"]/div/div[2]/div/div/ul/li/div/div[3]/div[2]/div/div[1]/div/div').text
                                            trCaseContent.append(dynamic)
                                        else:
                                            trCaseContent.append('--')
                                        # 点击 收起用例内容
                                        self.driver.find_element_by_xpath(
                                            '//*[@id="drawer-sidebar-workitemDetail"]/../div').click()
                                        time.sleep(1)
                                        # tr.click()
                                        for iLine in range(1, len(trCaseContent) + 1):
                                            sheet.cell(lineXLS, iLine,
                                                       str(trCaseContent[iLine - 1]))  # 将指定值写入第i+1行第j+1列
                                        # print(f'用例为{trCaseContent}')
                                        print(f'用例数量已增加为：{trCaseNum}')
                                        print(f'excel行数已增加为：{lineXLS}')

                                        if iTr >= 15:
                                            # 如果显示用例小于模块用例遇到未通过和暂缓，break（因为xpath会动态变化）
                                            if trNum < int(modelNameNum):
                                                if trCaseNum < int(modelNameNum):
                                                    break
                                    else:
                                        # tr.click()
                                        for iLine in range(1, len(trCaseContent) + 1):
                                            sheet.cell(lineXLS, iLine,
                                                       str(trCaseContent[iLine - 1]))  # 将指定值写入第i+1行第j+1列
                                        # print(f'用例为{trCaseContent}')
                                        print(f'当前模块用例数量已增加为：{trCaseNum}')
                                        print(f'excel行数已增加为：{lineXLS}')
                                        #  如果显示用例小于模块用例遇到最后一个需要滚动滑块
                                        if trNum < int(modelNameNum):
                                            if trCaseNum < int(modelNameNum):
                                                if iTr == trNum:
                                                    scrollbarXpath = tbodyXpath + f'/tr[{iTr}]' + '/td[1]/div/label/span[1]/input'
                                                    # print(scrollbarXpath)
                                                    scrollbar = self.driver.find_element_by_xpath(scrollbarXpath)
                                                    scrollbar.click()
                                                    scrollbar.send_keys(Keys.PAGE_DOWN)
                                else:
                                    print('用例已存在，跳过！！')
                            time.sleep(1)
                    else:
                        # 点击 选择用例
                        # 获取用例框
                        try:
                            tbodyXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody'
                            WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable((By.XPATH, tbodyXpath)))
                            tbody = self.driver.find_element_by_xpath(tbodyXpath)
                        except:
                            tbodyXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/div/div/table/tbody'
                            WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable((By.XPATH, tbodyXpath)))
                            tbody = self.driver.find_element_by_xpath(tbodyXpath)
                        trList = tbody.find_elements_by_xpath('tr')
                        trNum = len(trList)
                        print('***********用例模块中显示的用例数量:%s个**********' % trNum)
                        for iTr in range(1, trNum + 1):
                            trXpath = tbodyXpath + f'/tr[{iTr}]'
                            tr = self.driver.find_element_by_xpath(trXpath)
                            trContent = tr.text.split('\n')
                            TrCase = tr.text.replace('\n', ' ').replace('\r', ' ')  # 替换换行符号
                            # print(f'用例名称：{TrCase}')
                            print(f'用例变形前：{trContent}')
                            # print(f'用例列表：{trCaseList}')
                            bugTag = trContent[4]
                            if TrCase not in trCaseList:
                                trCaseNum += 1
                                lineXLS += 1
                                trCaseList.append(TrCase)
                                if '缺陷' not in trContent[4]:
                                    trContent.insert(4, '--')
                                else:
                                    trContent[4] = '是'
                                trCaseContent = trContent[0:2] + trContent[5:8] + trContent[2:5]
                                # print(f'变形后列表：{trCaseContent}')
                                if trCaseContent[6] == '未通过' or trCaseContent[6] == '暂缓':
                                    if '缺陷' in bugTag:
                                        tdBugXpath = trXpath + '/td[6]/div'
                                        # ---*****增加 将元素移到可视区域*****---
                                        tdBug = self.driver.find_element_by_xpath(tdBugXpath)
                                        self.driver.execute_script("arguments[0].scrollIntoView();", tdBug)
                                        tdBug.click()

                                        # self.driver.find_element_by_xpath(tdBugXpath).click()
                                        WebDriverWait(self.driver, 15).until(
                                            EC.element_to_be_clickable(
                                                (By.XPATH, '// *[text() = "缺陷标题"]')))  # 等待缺陷标题出现
                                        try:
                                            bugBodyXpath = '//*[text()="缺陷标题"]/../../../../../../../div[2]/table/tbody'
                                            bugBody = self.driver.find_element_by_xpath(bugBodyXpath)
                                            bugNum = bugBody.find_elements_by_xpath('tr')
                                        except:
                                            bugBodyXpath = '/html/body/div[7]/div/div/div[2]/div[2]/table/tbody'
                                            bugBody = self.driver.find_element_by_xpath(bugBodyXpath)
                                            bugNum = bugBody.find_elements_by_xpath('tr')
                                        BugId = ''
                                        for bugN in range(1, len(bugNum) + 1):
                                            tdBug.click()

                                            # self.driver.find_element_by_xpath(tdBugXpath).click()
                                            WebDriverWait(self.driver, 15).until(
                                                EC.element_to_be_clickable(
                                                    (By.XPATH, '// *[text() = "缺陷标题"]')))  # 等待缺陷标题出现
                                            bugNXpath = bugBodyXpath + f'/tr[{bugN}]/td[1]'
                                            bugN = self.driver.find_element_by_xpath(bugNXpath)
                                            # ---*****增加 将元素移到可视区域*****---
                                            self.driver.execute_script("arguments[0].scrollIntoView();", bugN)
                                            bugN.click()
                                            time.sleep(1)
                                            # 切换进入新窗口
                                            self.driver.switch_to.window(self.driver.window_handles[-1])
                                            WebDriverWait(self.driver, 15).until(
                                                EC.element_to_be_clickable(
                                                    (By.XPATH, '//*[contains(text(),"首页")]')))  # 等待页面加载
                                            print('判断权限')
                                            Test.getExistanceByxpath(xpath='//*[text()="暂无权限"]')
                                            # 判断是否有权限
                                            if element_exist == True:
                                                print('暂无权限')
                                                BugId = '暂无权限'
                                                time.sleep(1)
                                                self.driver.close()
                                                # 切换回原来窗口
                                                self.driver.switch_to.window(self.driver.window_handles[0])
                                            else:
                                                print('有权限')
                                                print(f'当前页面title为：{self.driver.title}')
                                                WebDriverWait(self.driver, 15).until(
                                                    EC.element_to_be_clickable(
                                                        (By.XPATH, '//*[text()="问题单号"]')))  # 等待问题单号出现
                                                BugValue = self.driver.find_element_by_xpath(
                                                    '//*[text()="问题单号"]/../../div[2]/div/div/span/input').get_attribute(
                                                    'value')
                                                print(f'问题单号为：{BugValue}')
                                                BugId = BugId + BugValue + '\n'
                                                time.sleep(1)
                                                self.driver.close()
                                                # 切换回原来窗口
                                                self.driver.switch_to.window(self.driver.window_handles[0])
                                        trCaseContent.append(BugId)
                                    else:
                                        trCaseContent.append('--')
                                    # 此处点击进入用例无需优化，因为已经可点击缺陷，此处应该可点击
                                    tdXpath = trXpath + '/td[2]/div'

                                    self.driver.find_element_by_xpath(tdXpath).click()
                                    # print('进入用例描述')
                                    try:
                                        WebDriverWait(self.driver, 15).until(
                                            EC.element_to_be_clickable(
                                                (By.XPATH, '//*[text()="前置条件"]')))  # 等待进入用例描述
                                    except:
                                        # 刷新
                                        self.driver.refresh()
                                        WebDriverWait(self.driver, 15).until(
                                            EC.element_to_be_clickable(
                                                (By.XPATH, '//*[text()="前置条件"]')))  # 等待进入用例描述
                                    dynamicXpath = '//*[@id="workitemModules"]/div/div[2]/div/div/ul/li/div/div[3]/div[2]/div/div[1]/div/div'
                                    Test.getExistanceByxpath(xpath=dynamicXpath)
                                    # 判断是否有动态内容
                                    if element_exist == True:
                                        dynamic = self.driver.find_element_by_xpath(
                                            '//*[@id="workitemModules"]/div/div[2]/div/div/ul/li/div/div[3]/div[2]/div/div[1]/div/div').text
                                        trCaseContent.append(dynamic)
                                    else:
                                        trCaseContent.append('--')
                                    # 点击 收起用例内容
                                    self.driver.find_element_by_xpath(
                                        '//*[@id="drawer-sidebar-workitemDetail"]/../div').click()
                                    time.sleep(1)
                                    # tr.click()
                                    for iLine in range(1, len(trCaseContent) + 1):
                                        sheet.cell(lineXLS, iLine,
                                                   str(trCaseContent[iLine - 1]))  # 将指定值写入第i+1行第j+1列
                                    # print(f'用例为{trCaseContent}')
                                    print(f'用例数量已增加为：{trCaseNum}')
                                    print(f'excel行数已增加为：{lineXLS}')
                                    # if iTr >= 15:
                                    #  如果显示用例小于模块用例遇到未通过和暂缓，break（因为xpath会动态变化）
                                    # if trNum < int(modelNameNum):
                                    #     if trCaseNum < int(modelNameNum):
                                    #         break
                                else:
                                    # tr.click()
                                    for iLine in range(1, len(trCaseContent) + 1):
                                        sheet.cell(lineXLS, iLine,
                                                   str(trCaseContent[iLine - 1]))  # 将指定值写入第i+1行第j+1列
                                    # print(f'用例为{trCaseContent}')
                                    print(f'当前模块用例数量已增加为：{trCaseNum}')
                                    print(f'excel行数已增加为：{lineXLS}')
                                    #  如果显示用例小于模块用例遇到最后一个需要滚动滑块
                                    if trNum < int(modelNameNum):
                                        if trCaseNum < int(modelNameNum):
                                            if iTr == trNum:
                                                scrollbarXpath = tbodyXpath + f'/tr[{iTr}]' + '/td[1]/div/label/span[1]/input'
                                                # print(scrollbarXpath)
                                                scrollbar = self.driver.find_element_by_xpath(scrollbarXpath)
                                                scrollbar.click()
                                                scrollbar.send_keys(Keys.PAGE_DOWN)
                            else:
                                print('用例已存在，跳过！！')
                        time.sleep(1)
                    self.driver.find_element_by_xpath('//*[text()="清空"]/..').click()
                # 如果没有用例
                else:
                    time.sleep(1)
                    self.driver.find_element_by_xpath('//*[text()="清空"]/..').click()

                print('成功、待测试用例----->')
                time.sleep(1)
                # 点击 进行筛选
                try:
                    # 尝试寻找元素，如若没有找到则会抛出异常
                    self.driver.find_element_by_xpath(
                        '/html/body/div[2]/main/header/section/section/section/span[2]/button').click()
                    # print('第一种')
                except:
                    self.driver.find_element_by_xpath(
                        '/html/body/div[3]/main/header/section/section/section/span[2]/button').click()
                    # print('第二种')
                time.sleep(1)
                # 添加状态
                element = self.driver.find_element_by_xpath(
                    '//*[@id="container"]/main/section/section[2]/section/div/div[1]/div[1]/div/div/div[5]/div[3]').click()
                time.sleep(1.5)

                # 点击 已通过
                self.driver.find_element_by_xpath(
                    '//*[@id="aoneCommonSearchDropdown"]/div[1]/ul/li[3]/div/label/span/input').click()
                time.sleep(0.5)
                # 点击 待测试
                self.driver.find_element_by_xpath(
                    '//*[@id="aoneCommonSearchDropdown"]/div[1]/ul/li[1]/div/label/span/input').click()
                time.sleep(0.5)
                self.driver.find_element_by_xpath('//*[text()="确定"]/..').click()
                time.sleep(1)
                # 点击 过滤
                self.driver.find_element_by_xpath('//*[text()="过滤"]/./..').click()
                time.sleep(1)
                WebDriverWait(self.driver, 15).until(
                    EC.element_to_be_clickable((By.XPATH,
                                                '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div')))  # 等待显示 过滤结果
                # 循环获取成功用例
                while trCaseNum < int(modelNameNum):
                    # 获取用例框
                    try:
                        tbodyXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody'
                        WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable((By.XPATH, tbodyXpath)))
                        tbody = self.driver.find_element_by_xpath(tbodyXpath)
                    except:
                        tbodyXpath = '//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/div/div/table/tbody'
                        WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable((By.XPATH, tbodyXpath)))
                        tbody = self.driver.find_element_by_xpath(tbodyXpath)
                    trList = tbody.find_elements_by_xpath('tr')
                    trNum = len(trList)
                    print('***********用例模块中显示的用例数量:%s个**********' % trNum)
                    for iTr in range(1, trNum + 1):
                        trXpath = tbodyXpath + f'/tr[{iTr}]'
                        tr = self.driver.find_element_by_xpath(trXpath)
                        trContent = tr.text.split('\n')
                        TrCase = tr.text.replace('\n', ' ').replace('\r', ' ')  # 替换换行符号
                        # print(f'用例名称：{TrCase}')
                        print(f'用例变形前：{trContent}')
                        # print(f'用例列表：{trCaseList}')
                        bugTag = trContent[4]
                        if TrCase not in trCaseList:
                            trCaseNum += 1
                            lineXLS += 1
                            trCaseList.append(TrCase)
                            if '缺陷' not in trContent[4]:
                                trContent.insert(4, '--')
                            else:
                                trContent[4] = '是'
                            trCaseContent = trContent[0:2] + trContent[5:8] + trContent[2:5]
                            # print(f'变形后列表：{trCaseContent}')
                            if trCaseContent[6] == '未通过' or trCaseContent[6] == '暂缓':
                                if '缺陷' in bugTag:
                                    tdBugXpath = trXpath + '/td[6]/div'
                                    self.driver.find_element_by_xpath(tdBugXpath).click()
                                    time.sleep(1.5)
                                    try:
                                        bugBodyXpath = '//*[text()="缺陷标题"]/../../../../../../../div[2]/table/tbody'
                                        bugBody = self.driver.find_element_by_xpath(bugBodyXpath)
                                        bugNum = bugBody.find_elements_by_xpath('tr')
                                    except:
                                        bugBodyXpath = '/html/body/div[7]/div/div/div[2]/div[2]/table/tbody'
                                        bugBody = self.driver.find_element_by_xpath(bugBodyXpath)
                                        bugNum = bugBody.find_elements_by_xpath('tr')
                                    BugId = ''
                                    for bugN in range(1, len(bugNum) + 1):
                                        bugNXpath = bugBodyXpath + f'/tr[{bugN}]/td[1]'
                                        self.driver.find_element_by_xpath(bugNXpath).click()
                                        time.sleep(1)
                                        # 切换进入新窗口
                                        self.driver.switch_to.window(self.driver.window_handles[-1])
                                        print('判断权限')
                                        Test.getExistanceByxpath(xpath='//*[text()="暂无权限"]')
                                        # 判断是否有权限
                                        if element_exist == True:
                                            print('暂无权限')
                                            BugId = '暂无权限'
                                            time.sleep(1)
                                            self.driver.close()
                                            # 切换回原来窗口
                                            self.driver.switch_to.window(self.driver.window_handles[0])
                                        else:
                                            print('有权限')
                                            print(f'当前页面title为：{self.driver.title}')
                                            WebDriverWait(self.driver, 15).until(
                                                EC.element_to_be_clickable((By.XPATH, '//*[text()="问题单号"]')))  # 等待问题单号出现
                                            BugValue = self.driver.find_element_by_xpath(
                                                '//*[text()="问题单号"]/../../div[2]/div/div/span/input').get_attribute('value')
                                            print(f'问题单号为：{BugValue}')
                                            BugId = BugId + BugValue + '\n'
                                            time.sleep(1)
                                            self.driver.close()
                                            # 切换回原来窗口
                                            self.driver.switch_to.window(self.driver.window_handles[0])
                                    trCaseContent.append(BugId)
                                else:
                                    trCaseContent.append('--')
                                tdXpath = trXpath + '/td[2]/div'
                                # ---*****增加 将元素移到可视区域*****---
                                tdBT = self.driver.find_element_by_xpath(tdXpath)
                                self.driver.execute_script("arguments[0].scrollIntoView();", tdBT)
                                tdBT.click()

                                # self.driver.find_element_by_xpath(tdXpath).click()
                                # print('进入用例描述')
                                try:
                                    WebDriverWait(self.driver, 15).until(
                                        EC.element_to_be_clickable((By.XPATH, '//*[text()="前置条件"]')))  # 等待进入用例描述
                                except:
                                    # 刷新
                                    self.driver.refresh()
                                    WebDriverWait(self.driver, 15).until(
                                        EC.element_to_be_clickable((By.XPATH, '//*[text()="前置条件"]')))  # 等待进入用例描述
                                dynamicXpath = '//*[@id="workitemModules"]/div/div[2]/div/div/ul/li/div/div[3]/div[2]/div/div[1]/div/div'
                                Test.getExistanceByxpath(xpath=dynamicXpath)
                                # 判断是否有动态内容
                                if element_exist == True:
                                    dynamic = self.driver.find_element_by_xpath(
                                        '//*[@id="workitemModules"]/div/div[2]/div/div/ul/li/div/div[3]/div[2]/div/div[1]/div/div').text
                                    trCaseContent.append(dynamic)
                                else:
                                    trCaseContent.append('--')
                                # 点击 收起用例内容
                                self.driver.find_element_by_xpath('//*[@id="drawer-sidebar-workitemDetail"]/../div').click()
                                time.sleep(1)
                                # tr.click()
                                for iLine in range(1,len(trCaseContent)+1):
                                    sheet.cell(lineXLS, iLine, str(trCaseContent[iLine-1]))  # 将指定值写入第i+1行第j+1列
                                # print(f'用例为{trCaseContent}')
                                print(f'用例数量已增加为：{trCaseNum}')
                                print(f'excel行数已增加为：{lineXLS}')
                                if iTr >= 15:
                                 # 如果显示用例小于模块用例遇到未通过和暂缓，break（因为xpath会动态变化）
                                    if trNum < int(modelNameNum):
                                        if trCaseNum < int(modelNameNum):
                                            break
                            else:
                                # tr.click()
                                for iLine in range(1,len(trCaseContent)+1):
                                    sheet.cell(lineXLS, iLine, str(trCaseContent[iLine-1]))  # 将指定值写入第i+1行第j+1列
                                # sheet.range(f'A{lineXLS}').value = trCaseContent
                                # print(f'用例为{trCaseContent}')
                                print(f'当前模块用例数量已增加为：{trCaseNum}')
                                print(f'excel行数已增加为：{lineXLS}')
                                #  如果显示用例小于模块用例遇到最后一个需要滚动滑块
                                if trNum < int(modelNameNum):
                                    if trCaseNum < int(modelNameNum):
                                        if iTr == trNum:
                                            scrollbarXpath = tbodyXpath + f'/tr[{iTr}]' + '/td[1]/div/label/span[1]/input'
                                            # print(scrollbarXpath)
                                            scrollbar = self.driver.find_element_by_xpath(scrollbarXpath)
                                            scrollbar.click()
                                            scrollbar.send_keys(Keys.PAGE_DOWN)
                        else:
                            print('用例已存在，跳过！！')
                self.driver.find_element_by_xpath('//*[text()="清空"]/..').click()
        return lineXLS

    '''
       函数名：getModel
       说明：获取测试计划所有模块
       参数：
       返回：
       作者：louwujian
    '''
    # 用例模块'//*[@id="container"]/main/section/section[2]/section/div/div[2]/div[2]/div/div/ul/div[2]/div/div[1]/div/nav/ul'
    def getModel(self, ulXpath, lineXLS, sheet):
        topModel = self.driver.find_element_by_xpath(ulXpath)
        liNumber = topModel.find_elements_by_xpath('li')
        # print(liNumber)
        # print(len(liNumber))
        if len(liNumber) > 0:
            # if len(liNumber) == 1:
            for iLi in range(1, len(liNumber) + 1):
                ul2Xpath =  ulXpath + f'/li[{iLi}]'
                top2Model = self.driver.find_element_by_xpath(ul2Xpath)
                ul2Number = top2Model.find_elements_by_xpath('ul')
                if len(ul2Number) == 1:
                    ul3Xpath = ul2Xpath + '/ul'
                    top3Model = self.driver.find_element_by_xpath(ul3Xpath)
                    li3Number = top3Model.find_elements_by_xpath('li')
                    for li3 in range(1, len(li3Number)+1):
                        li3Xpath = ul3Xpath + f'/li[{li3}]'
                        top4Model = self.driver.find_element_by_xpath(li3Xpath)
                        ul4Number = top4Model.find_elements_by_xpath('ul')
                        if len(ul4Number)==1:
                            ul4Xpath = li3Xpath + '/ul'
                            top5Model = self.driver.find_element_by_xpath(ul4Xpath)
                            li5Number = top5Model.find_elements_by_xpath('li')
                            for iLi5 in range(1, len(li5Number)+1):
                                li5Xpath = ul4Xpath + f'/li[{iLi5}]'
                                top6Model = self.driver.find_element_by_xpath(li5Xpath)
                                ul5Number = top6Model.find_elements_by_xpath('ul')
                                if len(ul5Number) == 1:
                                    ul5Xpath = li5Xpath + '/ul'
                                    top7Model = self.driver.find_element_by_xpath(ul5Xpath)
                                    li6Number = top7Model.find_elements_by_xpath('li')
                                    for iLi6 in range(1, len(li6Number)+1):
                                        li6Xpath = ul5Xpath + f'/li[{iLi6}]'
                                        top8Model = self.driver.find_element_by_xpath(li6Xpath)
                                        ul6Number = top8Model.find_elements_by_xpath('ul')
                                        if len(ul6Number) == 1:
                                            # print(f'有模块数：{len(ul6Number)}')
                                            ul6Xpath = li6Xpath + '/ul'
                                            top9Model = self.driver.find_element_by_xpath(ul6Xpath)
                                            li7Number = top9Model.find_elements_by_xpath('li')
                                            # print(f'有li数量：{len(li7Number)}')
                                            for iLi7 in range(1, len(li7Number)+1):
                                                li7Xpath = ul6Xpath + f'/li[{iLi7}]'
                                                top10Model = self.driver.find_element_by_xpath(li7Xpath)

                                                ul7Number = top10Model.find_elements_by_xpath('ul')
                                                if len(ul7Number) == 1:
                                                    # print(f'有模块数：{len(ul6Number)}')
                                                    ul7Xpath = li7Xpath + '/ul'
                                                    top11Model = self.driver.find_element_by_xpath(ul7Xpath)
                                                    li8Number = top11Model.find_elements_by_xpath('li')
                                                    # print(f'有li数量：{len(li7Number)}')
                                                    for iLi8 in range(1, len(li8Number) + 1):
                                                        li8Xpath = ul7Xpath + f'/li[{iLi8}]'
                                                        top12Model = self.driver.find_element_by_xpath(li8Xpath)

                                                        ul8Number = top12Model.find_elements_by_xpath('ul')
                                                        if len(ul8Number) == 1:
                                                            # print(f'有模块数：{len(ul6Number)}')
                                                            ul8Xpath = li8Xpath + '/ul'
                                                            top13Model = self.driver.find_element_by_xpath(ul8Xpath)
                                                            li9Number = top13Model.find_elements_by_xpath('li')
                                                            # print(f'有li数量：{len(li7Number)}')
                                                            for iLi9 in range(1, len(li9Number) + 1):
                                                                li9Xpath = ul8Xpath + f'/li[{iLi9}]'
                                                                top14Model = self.driver.find_element_by_xpath(li9Xpath)

                                                                ul9Number = top14Model.find_elements_by_xpath('ul')
                                                                if len(ul9Number) == 1:
                                                                    # print(f'有模块数：{len(ul6Number)}')
                                                                    ul9Xpath = li9Xpath + '/ul'
                                                                    top15Model = self.driver.find_element_by_xpath(
                                                                        ul9Xpath)
                                                                    li10Number = top15Model.find_elements_by_xpath('li')
                                                                    # print(f'有li数量：{len(li7Number)}')
                                                                    for iLi10 in range(1, len(li10Number) + 1):
                                                                        li10Xpath = ul8Xpath + f'/li[{iLi10}]'
                                                                        top16Model = self.driver.find_element_by_xpath(
                                                                            li10Xpath)
                                                                        modelName = top16Model.text
                                                                        modelNameNum = modelName.split('\n')[-1]
                                                                        print(f'此目录模块为：{modelName}')
                                                                        print('---------------')
                                                                        # 选择模块
                                                                        # ---*****增加 将元素移到可视区域*****---
                                                                        self.driver.execute_script(
                                                                            "arguments[0].scrollIntoView();", top16Model)
                                                                        top16Model.click()
                                                                        lineXLS = Test.getModelCase(lineXLS, sheet,
                                                                                                    modelNameNum)
                                                                else:
                                                                    modelName = top14Model.text
                                                                    modelNameNum = modelName.split('\n')[-1]
                                                                    print(f'此目录模块为：{modelName}')
                                                                    print('---------------')
                                                                    # 选择模块
                                                                    # ---*****增加 将元素移到可视区域*****---
                                                                    self.driver.execute_script(
                                                                        "arguments[0].scrollIntoView();", top14Model)
                                                                    top14Model.click()
                                                                    lineXLS = Test.getModelCase(lineXLS, sheet,
                                                                                                modelNameNum)
                                                        else:
                                                            modelName = top12Model.text
                                                            modelNameNum = modelName.split('\n')[-1]
                                                            print(f'此目录模块为：{modelName}')
                                                            print('---------------')
                                                            # 选择模块
                                                            # ---*****增加 将元素移到可视区域*****---
                                                            self.driver.execute_script(
                                                                "arguments[0].scrollIntoView();", top12Model)
                                                            top12Model.click()
                                                            lineXLS = Test.getModelCase(lineXLS, sheet, modelNameNum)
                                                else:
                                                    modelName = top10Model.text
                                                    modelNameNum = modelName.split('\n')[-1]
                                                    print(f'此目录模块为：{modelName}')
                                                    print('---------------')
                                                    # 选择模块
                                                    # ---*****增加 将元素移到可视区域*****---
                                                    self.driver.execute_script(
                                                        "arguments[0].scrollIntoView();", top10Model)
                                                    top10Model.click()
                                                    lineXLS = Test.getModelCase(lineXLS, sheet, modelNameNum)
                                        else:
                                            modelName = top8Model.text
                                            modelNameNum = modelName.split('\n')[-1]
                                            print(f'此目录模块为：{modelName}')
                                            print('---------------')
                                            # 选择模块
                                            # ---*****增加 将元素移到可视区域*****---
                                            self.driver.execute_script(
                                                "arguments[0].scrollIntoView();", top8Model)
                                            top8Model.click()
                                            lineXLS =Test.getModelCase(lineXLS, sheet, modelNameNum)
                                else:
                                    modelName = top6Model.text
                                    modelNameNum = modelName.split('\n')[-1]
                                    print(f'此目录模块为：{modelName}')
                                    print('---------------')
                                    # ---*****增加 将元素移到可视区域*****---
                                    self.driver.execute_script(
                                        "arguments[0].scrollIntoView();", top6Model)
                                    top6Model.click()
                                    lineXLS =Test.getModelCase(lineXLS, sheet, modelNameNum)
                        else:
                            modelName = top4Model.text
                            modelNameNum = modelName.split('\n')[-1]
                            print(f'此目录模块为：{modelName}')
                            print('---------------')
                            # 选择模块
                            # ---*****增加 将元素移到可视区域*****---
                            self.driver.execute_script(
                                "arguments[0].scrollIntoView();", top4Model)
                            top4Model.click()
                            lineXLS =Test.getModelCase(lineXLS, sheet, modelNameNum)
                else:
                    print('没有2级别目录')
        else:
            print('未找到一级目录')
        return lineXLS

    '''
       函数名：getCase
       说明：获取测试计划中的用例 主函数
       参数：
       返回：
       作者：louwujian
    '''
    def getCase(self, fileName = 'testCase'):
        # 新建一个表格文件
        book = Workbook()
        sheet = book.active
        # 获取用例数量
        WebDriverWait(self.driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '//div[text()="全部用例"]')))  # 等待显示 全部用例
        caseNum = self.driver.find_element_by_xpath('//div[text()="全部用例"]/../../span[2]').text
        print('测试计划包括用例: %s个' % caseNum)
        # 表头设置
        # print('开始设置表头')
        self.driver.find_element_by_xpath(
            '/html/body/div[3]/main/section/section[2]/section/div/div[2]/div[1]/div/button').click()
        time.sleep(1)
        # 表头-创建时间(去除)
        testCaseNum = Test.getHeaderAttr(xpath='//div[text()="创建时间"]/../div[2]', attr='aria-checked')
        if testCaseNum == 'true':
            self.driver.find_element_by_xpath('//div[text()="创建时间"]/../div[2]').click()
        # 表头-优先级（去除）
        time.sleep(1)
        testCaseNum = Test.getHeaderAttr(xpath='//div[text()="优先级"]/../div[2]', attr='aria-checked')
        if testCaseNum == 'true':
            self.driver.find_element_by_xpath('//div[text()="优先级"]/../div[2]').click()
        time.sleep(1)
        # 表头-测试用例编号（勾选）
        testCaseNum = Test.getHeaderAttr(xpath='//div[text()="测试用例编号"]/../div[2]', attr='aria-checked')
        if testCaseNum == 'false':
            self.driver.find_element_by_xpath('//div[text()="测试用例编号"]/../div[2]').click()
        # 表头-重要级别（勾选）
        time.sleep(1)
        testCaseNum = Test.getHeaderAttr(xpath='//div[text()="重要级别"]/../div[2]', attr='aria-checked')
        if testCaseNum == 'false':
            self.driver.find_element_by_xpath('//div[text()="重要级别"]/../div[2]').click()
        # 表头-自动化类型（勾选）
        time.sleep(1)
        testCaseNum = Test.getHeaderAttr(xpath='//div[text()="自动化类型"]/../div[2]', attr='aria-checked')
        if testCaseNum == 'false':
            self.driver.find_element_by_xpath('//div[text()="自动化类型"]/../div[2]').click()
        time.sleep(1)
        # 关闭表头设置
        self.driver.find_element_by_xpath('//span[text()="表头设置"]/../i').click()
        # 获取表头
        caseHeader = self.driver.find_element_by_xpath('/html/body/div[3]/main/section/section[2]/section/div/div[2]/div[1]/div').text
        # print(caseHeader)
        col = caseHeader.split('\n')
        if len(col) == 8:
            print('表头数量正常！')
        else:
            print(f'表头数量为：{len(col)} 个')
            print('表头数量不正确，导出数据会有异常，请重试或者检查系统分辨率！ 参考：1920×1080及以上（分辨率），100%（缩放比例）')
            return False
        colList = col[0:2] + col[5:8] + col[2:5] + ['问题单号', '备注']
        colList[7] = '提单'
        # print(colList)
        global lineXLS
        lineXLS = 1
        sum = 0
        # excel写入列名
        for c in range(1,len(colList)+1):
            sheet.cell(1, c, colList[c-1])  # 在第0行写入列名
        # 展开全部用例
        # expendXpath = '//*[@id="container"]/main/section/section[1]/div[2]/div/div/div[2]/div/div/button'
        expendTotalXpath = '//*[@id="container"]/main/section/section[1]/div[2]/div/div/div[2]'
        expendTotal = self.driver.find_element_by_xpath(expendTotalXpath)
        expendTotalNumber = expendTotal.find_elements_by_xpath('div')
        # 判断是否有多个产品用例库
        if len(expendTotalNumber) > 0:
            for iTotal in range (1, len(expendTotalNumber)+1):
                # 用例数
                caseNumAll = self.driver.find_element_by_xpath(f'(//div[text()="全部用例"])[{iTotal}]/../../span[2]').text
                expendModelXpath = expendTotalXpath + f'/div[{iTotal}]'
                expendXpath = expendModelXpath + f'/div/button'
                # print('展开所有用例模块')
                self.driver.find_element_by_xpath(expendXpath).click()  # 点击展开所有用例
                time.sleep(1)
                topHeaderXpath = expendTotalXpath + f'/div/ul/li/ul'
                # topHeaderXpath = '//*[@id="container"]/main/section/section[1]/div[2]/div/div/div[2]/div/ul/li/ul'
                lineXLS = Test.getModel(topHeaderXpath, lineXLS, sheet)
                sum = lineXLS- 1 -sum
                print('*'*30)
                print(f'第{iTotal}个产品，用例总数:{sum}')
                if str(sum) == str(caseNumAll):
                    print(f'第{iTotal}个产品，测试用例数量输出正确,校验通过！')
                else:
                    print(f'第{iTotal}个产品，测试用例数量输出异常############')
                print('*' * 30)
        else:
            print('未找到测试用例库')
        try:
            save_path = f'./output/planCaseAuto/{fileName} 测试结果记录表.xlsx'
            book.save(save_path)
            # 保存并关闭新文件
            print(f'输出文件保存正常：{save_path}')
        except:
            save_path = f'./output/planCaseAuto/{fileName} 测试结果记录表-副本.xlsx'
            book.save(save_path)
            print(f'文件保存异常，保存为副本：{save_path}')

    '''
       函数名：selectCaseMode
       说明：测试计划的用例挑选 主函数
       参数：selectType:云效编号-yunXiaoNum，测试用例编号-caseNum
       返回：
       作者：louwujian
    '''
    def selectCaseMode(self, tsPlanName = None, selectFile= selectCaseFileName, library=yxConfig.yxProductName, selectType='yunXiaoNum'):
        # 打开挑选用例表格文件
        # 获取工作簿对象
        wb = openpyxl.load_workbook(selectFile, data_only=True)
        sheet = wb.worksheets[0]
        print(f'sheet页名称为：{sheet.title}；参考的测试计划列名为：{tsPlanName}')
        # 获取工作表总行数
        rows = sheet.max_row
        # 获取工作表总列数
        cols = sheet.max_column
        # 读取第一行的所有内容
        row_list = []
        caseNum = 0
        planNum = 0
        cellName = '测试用例编号'
        if selectType =='yunXiaoNum':
            cellName ='编号'
        elif selectType =='caseNum':
            cellName ='测试用例编号'
        for i in range(1, cols + 1):
            cell_value = sheet.cell(row=1, column=i).value
            row_list.append(cell_value)
            if cell_value == cellName:
                caseNum = i
            if tsPlanName in cell_value:
                planNum = i
        if caseNum == 0:
            print(f'表头无测试用例编号，请检查 yxConfig.py 数据 或者《{yxConfig.selectCaseFileName}》 数据！')
            return False
        if planNum == 0:
            print(f'表头无匹配的测试计划名称，请检查 yxConfig.py 数据 或者《{yxConfig.selectCaseFileName}》 数据！')
            return False
        # print(row_list)
        # 读取第 X 列的所有用例编号并判断是否需要引入
        column_list = []
        for i in range(2, rows + 1):
            cell_value = sheet.cell(row=i, column=caseNum).value
            cellPlan_value = sheet.cell(row=i, column=planNum).value
            # if ('是' in cellPlan_value) and (cell_value is not None):
            if '是' in str(cellPlan_value) and (cell_value is not None):
                column_list.append(cell_value)
        # print(column_list)
        print(f'计划添加的用例数量为： {len(column_list)} 个')
        try:
            WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable(
                    (By.XPATH, '//*[@class="teamix-title teamix-title-tooltip-trigger"]')))  # 等待显示测试计划title
            planTitle = self.driver.find_element_by_xpath(
                '//*[@class="teamix-title teamix-title-tooltip-trigger"]').text
        except:
            WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@class="teamix-title"]')))  # 等待显示测试计划title
            planTitle = self.driver.find_element_by_xpath('//*[@class="teamix-title"]').text
        print(planTitle)
        if planTitle in tsPlanName:
            print('测试计划选择正确！')
        # 点击添加用例
        self.driver.find_elements(By.XPATH, ('//span[text()="添加用例"]'))[0].click()
        WebDriverWait(self.driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@class="next-dialog-header"]')))  # 等待显示 添加用例框
        # 选择用例库
        self.driver.find_element_by_xpath('//*[@class="next-dialog-header"]/../div[2]/div/div[1]/div[1]/span').click()
        time.sleep(1)
        libraryXpath = '//div[text()="' + library + '"]'
        print(f'选择的产品用例库为：{library}')
        # print(libraryXpath)
        WebDriverWait(self.driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@role="listbox"]')))  # 等待显示 用例库名称
        self.driver.find_element_by_xpath(libraryXpath).click()
        time.sleep(2)
        # 点击 全部用例
        try:
            print('点击全部用例2')
            allCasebtXpath = '(//div[text()="全部用例"])[2]'
            self.driver.find_element_by_xpath(allCasebtXpath).click()
        except:
            print('点击全部用例1')
            allCasebtXpath = '(//div[text()="全部用例"])[1]'
            self.driver.find_element_by_xpath(allCasebtXpath).click()
        time.sleep(2)
        # 点击 全部用例
        try:
            print('点击全部用例2')
            allCasebtXpath = '(//div[text()="全部用例"])[2]'
            self.driver.find_element_by_xpath(allCasebtXpath).click()
        except:
            print('点击全部用例1')
            allCasebtXpath = '(//div[text()="全部用例"])[1]'
            self.driver.find_element_by_xpath(allCasebtXpath).click()
        WebDriverWait(self.driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@class="next-dialog-header"]/../div[2]/div/div[2]/div/div[2]/div/div/ul/div[1]')))  # 等待显示 所有用例
        # 判断添加用例模式
        if selectType =='yunXiaoNum':
            caseInput = self.driver.find_element_by_xpath(
                '//input[@placeholder="搜索标题或编号"]')
            caseInput.send_keys("云效编号")  # 输入内容
            caseInput.send_keys(Keys.ENTER)
        elif selectType =='caseNum':
            filterXpath = '//*[@class="next-dialog-header"]/../div[2]/div/div[2]/div/div[1]/div[2]/span[3]'
            self.driver.find_element_by_xpath(filterXpath).click()
            filterAttr = Test.getHeaderAttr(xpath=filterXpath, attr='aria-expanded')
            print(filterAttr)
            for iTime in range(10):
                time.sleep(1)
                if filterAttr == 'true':
                    self.driver.find_element_by_xpath('//*[text()="测试用例编号"]').click()
                    time.sleep(1)
                    break
                else:
                    print('未选择到测试用例编号搜索框')
            caseInput = self.driver.find_element_by_xpath(
                '//input[@placeholder="请输入文本"]')
            caseInput.send_keys("用例编号")  # 输入内容
            self.driver.find_element_by_xpath('//*[text()="确定"]').click()
        time.sleep(1)
        # 筛选后 中间区域显示的内容
        containTextXpath = '//*[@class="next-dialog-header"]/../div[2]/div/div[2]/div/div[2]/div/div'
        # WebDriverWait(self.driver, 15).until(
        #     EC.element_to_be_clickable((By.XPATH, containTextXpath)))  # 等待显示 用例
        # time.sleep(0.5)
        for TextItem in range(30):
            time.sleep(0.5)
            try:
                containText = self.driver.find_element_by_xpath(containTextXpath).text
                if containText != '':
                    break
            except:
                print('')
        # 遍历查找每个用例
        selectCaseErrorList = []
        haveSelectNum = 0
        NoCaseNum = 0
        for iCase in column_list:
            # print(iCase)
            for iTime in range(10):
                time.sleep(0.5)
                if selectType == 'yunXiaoNum':
                    for iValue in range(10):
                        time.sleep(0.5)
                        caseValueAttr = Test.getHeaderAttr(xpath='//input[@placeholder="搜索标题或编号"]', attr='value')
                        # print(caseValueAttr)
                        if caseValueAttr != iCase:
                            # print('未输入编号')
                            caseInput = self.driver.find_element_by_xpath(
                                '//input[@placeholder="搜索标题或编号"]')
                            caseInput.send_keys(Keys.CONTROL, "a")  # 相当于ctrl + a快捷键全选
                            caseInput.send_keys(Keys.DELETE)  # 快捷键删除
                            caseInput.send_keys(iCase)
                        else:
                            # print('编号输入正确！，点击确认！')
                            caseInput.send_keys(Keys.ENTER)
                            break
                    # time.sleep(0.5)
                    break
                elif selectType == 'caseNum':
                    caseNumAttr = Test.getHeaderAttr(xpath='//*[text()="测试用例编号"]/../..', attr='aria-expanded')
                    if caseNumAttr == 'true':
                        for iValue in range(10):
                            time.sleep(0.5)
                            caseValueAttr = Test.getHeaderAttr(xpath='//input[@placeholder="请输入文本"]', attr='value')
                            if caseValueAttr not in iCase:
                                caseInput = self.driver.find_element_by_xpath(
                                    '//input[@placeholder="请输入文本"]')
                                caseInput.send_keys(Keys.CONTROL, "a")  # 相当于ctrl + a快捷键全选
                                caseInput.send_keys(Keys.DELETE)  # 快捷键删除
                                caseInput.send_keys(iCase)
                            else:
                                # print('测试用例编号输入正确！，点击确认！')
                                if selectType == 'yunXiaoNum':
                                    caseInput.send_keys(Keys.ENTER)
                                elif selectType == 'caseNum':
                                    self.driver.find_element_by_xpath('//*[text()="确定"]').click()
                                break
                        # time.sleep(0.5)
                        break
                    else:
                        # print('未选择到测试用例编号输入框，需点击')
                        self.driver.find_element_by_xpath('//*[text()="测试用例编号"]').click()

            # 筛选后 中间区域显示的内容
            containTextXpath = '//*[@class="next-dialog-header"]/../div[2]/div/div[2]/div/div[2]/div/div'
            # WebDriverWait(self.driver, 15).until(
            #     EC.element_to_be_clickable((By.XPATH, containTextXpath)))  # 等待显示 用例
            # time.sleep(0.5)
            for TextItem in range(30):
                time.sleep(0.5)
                try:
                    containText = self.driver.find_element_by_xpath(containTextXpath).text
                    if containText != '':
                        break
                except:
                    print('')
            # print(containText)
            if '暂无内容' not in containText:
                # 获取筛选后的用例
                caseCheckXpath = '//*[@class="next-dialog-header"]/../div[2]/div/div[2]/div/div[2]'
                caseCheckModel = self.driver.find_element_by_xpath(caseCheckXpath)
                caseCheckNumber = caseCheckModel.find_elements_by_xpath('div')
                # 判断筛选后的用例数量
                print(len(caseCheckNumber))
                if len(caseCheckNumber) == 1:
                    try:
                        # 只有1个用例
                        checkedXpath = '//*[@class="next-dialog-header"]/../div[2]/div/div[2]/div/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody/tr/td[1]/div/label/span[1]/input'
                        checkedAttr = Test.getHeaderAttr(xpath=checkedXpath, attr='aria-checked')
                    except:
                        # 只有1个用例--特殊情况
                        checkedXpath = '/html/body/div[4]/div[2]/div/div/div[2]/div/div[2]/div/div[2]/div/div/ul/div[1]/div/div[2]/div/div[2]/div[2]/table/tbody/tr/td[1]/div/label/span[1]/input'
                        checkedAttr = Test.getHeaderAttr(xpath=checkedXpath, attr='aria-checked')
                if len(caseCheckNumber) > 1:
                    # 筛选出多个用例
                    trCheckXpath = '//*[@class="next-dialog-header"]/../div[2]/div/div[2]/div/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody/tr/td[3]'
                    trCheck_elements = self.driver.find_elements_by_xpath(trCheckXpath)
                    for itr in range(1, len(trCheck_elements)+1):
                        # print(itr)
                        itrContent = self.driver.find_element_by_xpath(f'({trCheckXpath})[{itr}]').text
                        # print(itrContent)
                        # 判断用例是否为需要的
                        if iCase == itrContent:
                            checkedXpath = f'(//*[@class="next-dialog-header"]/../div[2]/div/div[2]/div/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody/tr/td[1]/div/label/span[1]/input)[{itr}]'
                            checkedAttr = Test.getHeaderAttr(xpath=checkedXpath, attr='aria-checked')
                            break
                # try:
                #     checkedXpath = '//*[@class="next-dialog-header"]/../div[2]/div/div[2]/div/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody/tr/td[1]/div/label/span[1]/input'
                #     checkedAttr = Test.getHeaderAttr(xpath=checkedXpath, attr='aria-checked')
                # except:
                #     checkedXpath = '//*[@class="next-dialog-header"]/../div[2]/div/div[2]/div/div[2]/div/div/ul/div[1]/div/div[2]/div/div[2]/div[2]/table/tbody/tr/td[1]/div/label/span[1]/input'
                #     checkedAttr = Test.getHeaderAttr(xpath=checkedXpath, attr='aria-checked')
                if checkedAttr == 'false':
                    # 选择用例
                    print(f'{iCase} 正常')
                    self.driver.find_element_by_xpath(checkedXpath).click()
                else:
                    print(f'{iCase} 此用例已被选择（添加前已被选择 或者重复用例！）')
                    selectCaseErrorList.append(iCase + ',已被选择或者重复')
                    haveSelectNum +=1
            elif '暂无内容' in containText:
                print(f'{iCase} 未找到用例！')
                selectCaseErrorList.append(iCase + ',未找到用例')
                NoCaseNum +=1
            else:
                print('ERROR：自动退出或其他网络原因，请重新执行！')
                return False
        selectAllXpath = '//*[@class="next-dialog-header"]/../div[2]/div/div[2]/div/div[1]/div[3]/div[2]'
        selectAll = self.driver.find_element_by_xpath(selectAllXpath).text
        print(selectAll)
        print('*' * 30)
        print(f'计划添加的用例数量为： {len(column_list)} 条')
        print('')
        if '已选中' not in selectAll:
            print('已选中 0 条用例')
            selectAllNum = 0
            #
            # 已选择用例完成，但是没选中的，点击"取消"
            self.driver.find_element_by_xpath('//*[text()="取消"]').click()
        else:
            try:
                selectAllList = selectAll.split(' ')
                selectAllNum = int(selectAllList[1])
                print(selectAll)
            except:
                selectAllNum = int(re.findall(r'\d+', selectAll)[0])
            # 已选用例完成，点击"添加"
            self.driver.find_element_by_xpath('//*[text()="添加"]').click()
        print(f'之前已被选择或者重复用例 {haveSelectNum} 条')
        print(f'未找到用例 {NoCaseNum} 条')
        # print(f'已选{selectAllNum}')
        allNum = selectAllNum + haveSelectNum + NoCaseNum
        # print(allNum)
        print('*' * 30)
        if allNum == len(column_list):
            print('用例挑选总数正常！如有疑问可查看noSelectCase-XXX.text自检！')
        else:
            print('用例挑选总数异常，可联系开发人员反馈！')
        # 保存log
        path = './output/selectCaseAuto'
        if not os.path.exists(path):
            os.mkdir(path)
        pathFile = './output/selectCaseAuto/noSelectCase'
        current_datetime = datetime.now()
        str_current_datetime = current_datetime.strftime("%Y-%m-%d %H-%M-%S")  # 获取时间
        # print("Current date & time : ", str_current_datetime)
        file_name = pathFile + str_current_datetime + ".txt"
        with open(file_name, 'w', encoding='utf-8') as f:
            print(f'文件保存路径：{file_name}')
            checkStr = '\n'.join(selectCaseErrorList)
            f.write(checkStr)  # 内容写入未匹配到的用例编号
        f.close()

    '''
       函数名：editFile
       说明：测试计划的用例引用模式：编辑导出的用例、并筛选需引用的待测、暂缓、未通过项
       参数：fileNamePath:导出的测试结果路径及名称
       返回：
       作者：louwujian
    '''
    def editFile(self, fileNamePath='testFile'):
        # 获取工作簿对象
        wb = openpyxl.load_workbook(fileNamePath, data_only=True)

        sheet = wb.worksheets[0]
        print(f'sheet页名称为：{sheet.title}；参考的测试计划列名为：{fileNamePath}')
        # 保存excel
        path = './output/quoteCaseAuto'
        if not os.path.exists(path):
            os.mkdir(path)
        new_wb = wb.active
        ws_copy = wb.copy_worksheet(new_wb)
        pathFileName = f'./output/quoteCaseAuto/{autoPlanName} 用例引用记录表.xlsx'
        wb.save(pathFileName)
        print(f'输出的引用记录表为：{pathFileName}')

        # 打开挑选用例表格文件
        # 获取工作簿对象
        edit_wb = openpyxl.load_workbook(pathFileName, data_only=True)
        edit_sheet = edit_wb.worksheets[0]
        print(f'sheet页名称为：{edit_sheet.title}；参考的测试计划列名为：{pathFileName}')
        # 获取工作表总行数
        edit_rows = edit_sheet.max_row
        # print(edit_rows)
        # 获取工作表总列数
        edit_cols = edit_sheet.max_column
        # print(edit_cols)
        edit_sheet.cell(1, edit_cols+1, newPlanName)
        statusNum = 0
        # 读取第一行
        for j in range(1, edit_cols+1):
            cols_value = edit_sheet.cell(row=1, column=j).value
            if cols_value == '状态':
                statusNum = j
                print(statusNum)

        # 读取第 X 列的所有状态并判断是否需要引入
        column_list = []
        for i in range(2, edit_rows + 1):
            cells_value = edit_sheet.cell(row=i, column=statusNum).value
            if cells_value =='待测试'or cells_value == '未通过' or cells_value== '暂缓':
                edit_sheet.cell(i, edit_cols+1, '是')
        edit_wb.save(pathFileName)
        print(f'{pathFileName} 编辑完成！')
        #

        print("=========================测试引用模式下（用例挑选模式）================================")
        # 直接进入 测试总计划页面################
        uu1 = 'https://devops.aliyun.com/testhub/plan#viewIdentifier=fc0ddd9669fd618c2c9729e39b'
        self.driver.get(uu1)
        # print('直接进入 测试计划页面')
        autoPlanNameList = re.split(',|，', newPlanName)
        if len(autoPlanNameList) != 1:
            print('此模式测试计划数量只能有1个，请检查测试计划名称！')
            return
        print(f'需挑选用例的测试计划为：{autoPlanNameList}')
        testPlanNum = 0
        for planName in autoPlanNameList:
            testPlanNum += 1
            print(f'------开始第 {testPlanNum} 个测试计划------')
            self.driver.get(uu1)
            WebDriverWait(self.driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, '//*[text()="计划名称"]')))  # 等待显示 计划名称
            # print('直接进入 测试计划页面')
            plan = '//*[text()=\"' + planName + '\"]'
            print('测试计划为:%s' % planName)
            time.sleep(2)
            Test.getExistanceByxpath(xpath=plan)
            # 判断是否有测试计划
            if element_exist == True:
                print('当前页测试计划存在，可正常点击！')
            else:
                print('当前页测试计划不存在，需翻页！')
                # 配置每页数量
                pageXpath = '//*[text()="每页显示:"]/../span[2]'
                page = self.driver.find_element_by_xpath(pageXpath).click()
                try:
                    pageMAXXpath = '/html/body/div[5]/div/ul/li[3]/div/span'
                    pageMAX = self.driver.find_element_by_xpath(pageMAXXpath).click()
                except:
                    pageMAXXpath = '/html/body/div[4]/div/ul/li[3]/div/span'
                    pageMAX = self.driver.find_element_by_xpath(pageMAXXpath).click()
                WebDriverWait(self.driver, 15).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[text()="计划名称"]')))  # 等待显示 计划名称
                # 获取页数
                pageNumXpath = '//*[@id="container"]/div/main/section/section/section/div/div/div/div/div[2]/div[2]/div[1]/div'
                pageNumModel = self.driver.find_element_by_xpath(pageNumXpath)
                pageNum = pageNumModel.find_elements_by_xpath('button')
                for pNum in range(1, len(pageNum) + 1):
                    pageNumButtonXpath = pageNumXpath + f'/button[{pNum}]'
                    self.driver.find_element_by_xpath(pageNumButtonXpath).click()
                    WebDriverWait(self.driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[text()="计划名称"]')))  # 等待显示 计划名称
                    Test.getExistanceByxpath(xpath=plan)
                    # print(element_exist)
                    # 判断是否有测试计划
                    if element_exist == True:
                        print(f'进入第{pNum}页，测试计划存在，可正常点击！')
                        break
                    else:
                        print(f'进入第{pNum}页未找到测试计划！')
            # 选择计划
            self.driver.find_element_by_xpath(plan).click()
        Test.selectCaseMode(tsPlanName=newPlanName, selectFile=pathFileName)

    '''
       函数名：action
       说明： 执行函数，判断执行哪种标记
       参数：mode 执行方式 自动化类型标注 "autoType"  ，自动化结果标注"autoResult"
            method 自动化类型标注中产品选择 search 搜索选择，link 链接跳转（推荐）
            product 产品名称
       返回：
       作者：louwujian
    '''

    def action(self, mode=autoLabel, method='link', product=yxProductName, line=lineNum):
        # 自动化用例类型标注
        if mode == 'autoType':
            print("=========================自动化类型标注模式================================")
            caseList = self.readCsv(yxConfig.typeFileName)
            print(caseList)
            # 选择产品
            if method == 'link':
                time.sleep(2)
                # 直接进入对应产品 测试用例库页面################
                if product == '01_BIOS':
                    uu1 = 'https://devops.aliyun.com/testhub/repo/d232bebdbc785186a5af2a0b3d/case#selectedTreeIdentifier=162583896a8f4c41c10b4548e6'
                    self.driver.get(uu1)
                    print('直接进入 01_BIOS 测试用例页面')
                elif product == '02_BMC':
                    uu1 = 'https://devops.aliyun.com/testhub/repo/18ea8772d6eee2c3cc8a12f8fb/case#selectedTreeIdentifier=2e8400d2a1982e6a8dfc03c42b'
                    self.driver.get(uu1)
                    print('直接进入 02_BMC 测试用例页面')
                elif product == '03_昆仑卫士':
                    uu1 = 'https://devops.aliyun.com/testhub/repo/1ca7c1dfbea00fa106934b2fe6/case#selectedTreeIdentifier=b13004ccd65c9562d59acae9e1'
                    self.driver.get(uu1)
                    print('直接进入 03_昆仑卫士 测试用例页面')
                elif product == '04_机房管理系统':
                    uu1 = 'https://devops.aliyun.com/testhub/repo/d77968bfd8198dbfcf3562abf0/case#selectedTreeIdentifier=e65b7751396bfa0d0e791f6644'
                    self.driver.get(uu1)
                    print('直接进入 04_机房管理系统 测试用例页面')
            elif method == 'search':
                # 搜索产品
                self.driver.find_element_by_xpath('//span[text()="在 云效 中搜索"]').click()
                time.sleep(2)
                # 输入产品名称
                self.driver.find_element_by_xpath('//*[@id="yxGlobalSearchInput"]').send_keys(product)
                time.sleep(1)
                self.driver.find_element_by_xpath('//*[@id="yxGlobalSearchInput"]').send_keys(Keys.ENTER)
                # 切换进入新的窗口
                self.driver.switch_to.window(self.driver.window_handles[-1])
            time.sleep(5)
            # 全部标记
            if line == 'all':
                print("********************全部标记模式******************")
                m = 0
                for i in caseList:
                    m += 1
                    print('开始，标记第%d个用例' % m)
                    print(i)
                    iList = i.split(',')
                    listId = iList[0]
                    print(listId)
                    listType = iList[-1]
                    print(listType)
                    Test.labelAuto(CaseId=listId, CaseType=listType)
                    print('结束，标记第%d个用例' % m)

                self.driver.quit()
                print('结束关闭浏览器')
            # 标记所填行号
            else:
                print("********************部分标记模式******************")
                lineList = line.split(',')
                print(lineList)
                for num in lineList:
                    m = 0
                    for i in caseList:
                        m += 1
                        if num == str(m):
                            print('开始，标记第%d行的用例类型' % m)
                            print(i)
                            iList = i.split(',')
                            listId = iList[0]
                            print(listId)
                            listType = iList[-1]
                            print(listType)
                            Test.labelAuto(CaseId=listId, CaseType=listType)
                            print('结束，标记第%d行的用例类型' % m)

                self.driver.quit()
                # print('结束关闭浏览器')
            pathFile = './output/labelAuto/noLabelCase'
            current_datetime = datetime.now()
            str_current_datetime = current_datetime.strftime("%Y-%m-%d %H-%M-%S")  # 获取时间
            print("Current date & time : ", str_current_datetime)
            file_name = pathFile + str_current_datetime + ".txt"
            with open(file_name, 'w', encoding='utf-8') as f:
                print(f)
                checkStr = '\n'.join(noLabellist)
                f.write(checkStr)  # 内容写入未匹配到的用例编号
            f.close()

        # 自动化测试结果标注
        elif mode == 'autoResult':
            caseList = self.readCsv(yxConfig.resultFileName)
            print("=========================自动化结果标注模式================================")
            print(caseList)
            # 直接进入 测试总计划页面################
            uu1 = 'https://devops.aliyun.com/testhub/plan#viewIdentifier=fc0ddd9669fd618c2c9729e39b'
            self.driver.get(uu1)
            print('直接进入 测试计划页面')
            autoPlanNameList = re.split(',|，', autoPlanName)
            if len(autoPlanNameList) != 1:
                print('测试计划数量只能有1个，请检查测试计划名称！')
                return
            print(f'需标记的测试计划为：{autoPlanName}')
            plan = '//*[text()=\"' + autoPlanName + '\"]'
            time.sleep(2)
            # 选择计划
            self.driver.find_element_by_xpath(plan).click()
            time.sleep(2.5)
            # 点击 进行筛选
            self.driver.find_element_by_xpath('/html/body/div[3]/main/header/section/section/section/span[2]/button').click()
            time.sleep(2)
            self.driver.find_element_by_xpath('//*[text()="类型"]').click()
            time.sleep(2)
            self.driver.find_element_by_xpath('//*[text()="其他"]').click()
            time.sleep(2)
            self.driver.find_element_by_xpath('//*[text()="测试用例编号"]').click()
            time.sleep(2)
            # 点击 过滤
            self.driver.find_element_by_xpath('//*[text()="过滤"]/./..').click()
            # print('点击 过滤')
            time.sleep(3)
            m = 0
            for i in caseList:
                print('开始，标记第%d个用例' % m)
                print(i)
                # if m % 100 == 0:
                #     self.driver.refresh()
                #     time.sleep(3)
                iList = i.split(',')
                listId = iList[0]
                # print(listId)
                listResult= iList[-1]
                # print(listResult)
                try:
                    Test.resultAuto(CaseId=listId, CaseResult=listResult)
                except:
                    self.driver.refresh()
                    time.sleep(3)
                    try:
                        print('第一次出错，刷新，加入未标记列表******')
                        list2.append(listId + ',第一次出错')
                        #判断是否需收起用例
                        Test.getExistanceByxpath(xpath='//*[text()="前置条件"]')
                        if element_exist == True:
                            # 点击 收起用例内容
                            self.driver.find_element_by_xpath('//*[@id="drawer-sidebar-workitemDetail"]/../div').click()
                            print('收起用例')
                        #判断是否已退出登录
                        Test.getExistanceByxpath(xpath='//*[text()="账号登录"]')
                        if element_exist == True:
                            print('账号自动退出,需再次执行！！！！！！')
                            break
                        time.sleep(1)
                        Test.resultAuto(CaseId=listId, CaseResult=listResult)

                    except:
                        print('第二次出错，刷新，加入未标记列表******')
                        list2.append(listId + ',确认错误')
                print('结束，标记第%d个用例' % m)
                m += 1
            self.driver.quit()
            print('结束关闭浏览器')
            pathFile = './output/resultAuto/noResult'
            current_datetime = datetime.now()
            str_current_datetime = current_datetime.strftime("%Y-%m-%d %H-%M-%S")  # 获取时间
            print("Current date & time : ", str_current_datetime)
            file_name = pathFile + str_current_datetime + ".txt"
            with open(file_name, 'w', encoding='utf-8') as f:
                print(f)
                checkStr = '\n'.join(list2)
                f.write(checkStr)  # 内容写入未匹配到的用例编号
            f.close()

        # 测试计划用例导出模式
        elif mode == 'getPlanCase':
            print("=========================测试计划用例导出模式================================")
            # 直接进入 测试总计划页面################
            uu1 = 'https://devops.aliyun.com/testhub/plan#viewIdentifier=fc0ddd9669fd618c2c9729e39b'
            self.driver.get(uu1)
            # print('直接进入 测试计划页面')
            autoPlanNameList = re.split(',|，', autoPlanName)
            print(f'需获取测试计划数量：{len(autoPlanNameList)}')
            testPlanNum = 0
            for planName in autoPlanNameList:
                testPlanNum += 1
                print(f'------开始第 {testPlanNum} 个测试计划------')
                # 开始计时
                startTime = time.clock()
                self.driver.get(uu1)
                WebDriverWait(self.driver, 15).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[text()="计划名称"]')))  # 等待显示 计划名称
                # print('直接进入 测试计划页面')
                plan = '//*[text()=\"' + planName + '\"]'
                print('测试计划为:%s' % planName)
                time.sleep(2)
                Test.getExistanceByxpath(xpath=plan)
                # 判断是否有测试计划
                if element_exist == True:
                    print('当前页测试计划存在，可正常点击！')
                else:
                    print('当前页测试计划不存在，需翻页！')
                    # 配置每页数量
                    pageXpath = '//*[text()="每页显示:"]/../span[2]'
                    page = self.driver.find_element_by_xpath(pageXpath).click()
                    try:
                        pageMAXXpath = '/html/body/div[5]/div/ul/li[3]/div/span'
                        pageMAX = self.driver.find_element_by_xpath(pageMAXXpath).click()
                    except:
                        pageMAXXpath = '/html/body/div[4]/div/ul/li[3]/div/span'
                        pageMAX = self.driver.find_element_by_xpath(pageMAXXpath).click()
                    WebDriverWait(self.driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[text()="计划名称"]')))  # 等待显示 计划名称
                    # 获取页数
                    pageNumXpath = '//*[@id="container"]/div/main/section/section/section/div/div/div/div/div[2]/div[2]/div[1]/div'
                    pageNumModel = self.driver.find_element_by_xpath(pageNumXpath)
                    pageNum = pageNumModel.find_elements_by_xpath('button')
                    for pNum in range(1, len(pageNum)+1):
                        pageNumButtonXpath = pageNumXpath + f'/button[{pNum}]'
                        self.driver.find_element_by_xpath(pageNumButtonXpath).click()
                        WebDriverWait(self.driver, 15).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[text()="计划名称"]')))  # 等待显示 计划名称
                        Test.getExistanceByxpath(xpath=plan)
                        # print(element_exist)
                        # 判断是否有测试计划
                        if element_exist == True:
                            print(f'进入第{pNum}页，测试计划存在，可正常点击！')
                            break
                        else:
                            print(f'进入第{pNum}页未找到测试计划！')
                # 选择计划
                self.driver.find_element_by_xpath(plan).click()
                # time.sleep(3)
                Test.getCase(fileName=planName)
                # 当前计划结束，结束计时
                endTime = time.clock()
                # 计算用时
                print(f'测试导出计划：{planName},用时：{int(endTime-startTime)}秒')
            # 关闭浏览器
            self.driver.quit()
            print('结束关闭浏览器')

        # 测试用例挑选模式
        elif mode == 'selectCase':
            print("=========================测试用例挑选模式================================")
            # 直接进入 测试总计划页面################
            uu1 = 'https://devops.aliyun.com/testhub/plan#viewIdentifier=fc0ddd9669fd618c2c9729e39b'
            self.driver.get(uu1)
            # print('直接进入 测试计划页面')
            autoPlanNameList = re.split(',|，', autoPlanName)
            if len(autoPlanNameList) != 1:
                print('此模式测试计划数量只能有1个，请检查测试计划名称！')
                return
            print(f'需挑选用例的测试计划为：{autoPlanNameList}')
            testPlanNum = 0
            for planName in autoPlanNameList:
                testPlanNum += 1
                print(f'------开始第 {testPlanNum} 个测试计划------')
                self.driver.get(uu1)
                WebDriverWait(self.driver, 15).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[text()="计划名称"]')))  # 等待显示 计划名称
                # print('直接进入 测试计划页面')
                plan = '//*[text()=\"' + planName + '\"]'
                print('测试计划为:%s' % planName)
                time.sleep(2)
                Test.getExistanceByxpath(xpath=plan)
                # 判断是否有测试计划
                if element_exist == True:
                    print('当前页测试计划存在，可正常点击！')
                else:
                    print('当前页测试计划不存在，需翻页！')
                    # 配置每页数量
                    pageXpath = '//*[text()="每页显示:"]/../span[2]'
                    page = self.driver.find_element_by_xpath(pageXpath).click()
                    try:
                        pageMAXXpath = '/html/body/div[5]/div/ul/li[3]/div/span'
                        pageMAX = self.driver.find_element_by_xpath(pageMAXXpath).click()
                    except:
                        pageMAXXpath = '/html/body/div[4]/div/ul/li[3]/div/span'
                        pageMAX = self.driver.find_element_by_xpath(pageMAXXpath).click()
                    WebDriverWait(self.driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[text()="计划名称"]')))  # 等待显示 计划名称
                    # 获取页数
                    pageNumXpath = '//*[@id="container"]/div/main/section/section/section/div/div/div/div/div[2]/div[2]/div[1]/div'
                    pageNumModel = self.driver.find_element_by_xpath(pageNumXpath)
                    pageNum = pageNumModel.find_elements_by_xpath('button')
                    for pNum in range(1, len(pageNum) + 1):
                        pageNumButtonXpath = pageNumXpath + f'/button[{pNum}]'
                        self.driver.find_element_by_xpath(pageNumButtonXpath).click()
                        WebDriverWait(self.driver, 15).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[text()="计划名称"]')))  # 等待显示 计划名称
                        Test.getExistanceByxpath(xpath=plan)
                        # print(element_exist)
                        # 判断是否有测试计划
                        if element_exist == True:
                            print(f'进入第{pNum}页，测试计划存在，可正常点击！')
                            break
                        else:
                            print(f'进入第{pNum}页未找到测试计划！')
                # 选择计划
                self.driver.find_element_by_xpath(plan).click()
                # time.sleep(3)
                # 开始计时
                startTime = time.clock()
                Test.selectCaseMode(tsPlanName=planName)
                # try:
                #     Test.selectCaseMode(fileName=planName)
                # except:
                #     # 判断是否已退出登录
                #     Test.getExistanceByxpath(xpath='//*[text()="账号登录"]')
                #     if element_exist == True:
                #         print("*"*30)
                #         print('Error：账号自动退出,需再次执行！！！！！！')
                #         print('Error：账号自动退出,需再次执行！！！！！！')
                #         print('Error：账号自动退出,需再次执行！！！！！！')
                #         return False
                #     time.sleep(1)
                # 当前计划结束，结束计时
                endTime = time.clock()
                # 计算用时
                print(f'测试计划：{planName},挑选用例用时：{int(endTime - startTime)}秒')
        #   关闭浏览器
            time.sleep(15)
            self.driver.quit()
            print('结束,关闭浏览器')

        # 测试用例引用模式
        elif mode == 'quoteCase':
            print("=========================测试单暂缓、失败、待测测试用例引用模式================================")
            # 判断是否有导出文件
            saveQuoteFile = f'{autoPlanName} 测试结果记录表.xlsx'
            print(saveQuoteFile)
            excelFile = ''
            fileList = os.listdir('./output/planCaseAuto/')
            print(fileList)
            # 开始计时
            startTime = time.clock()
            for fileName in fileList:
                print(fileName)
                if saveQuoteFile in fileName:
                    print('ok')
                    excelFile = fileName
                    break
            if excelFile:
                print(f'找到了:{excelFile}')
            else:
                print("未找到测试结果导出文件，会先进行导出")
                print("------------------------开始用例引用中的用例导出模式---------------------------------")
                # 直接进入 测试总计划页面################
                uu1 = 'https://devops.aliyun.com/testhub/plan#viewIdentifier=fc0ddd9669fd618c2c9729e39b'
                self.driver.get(uu1)
                # print('直接进入 测试计划页面')
                autoPlanNameList = re.split(',|，', autoPlanName)
                if len(autoPlanNameList) != 1:
                    print('此模式测试计划数量只能有1个，请检查测试计划名称！')
                    return
                print(f'需挑选用例的测试计划为：{autoPlanNameList}')
                testPlanNum = 0
                for planName in autoPlanNameList:
                    testPlanNum += 1
                    print(f'------开始第 {testPlanNum} 个测试计划------')

                    self.driver.get(uu1)
                    WebDriverWait(self.driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[text()="计划名称"]')))  # 等待显示 计划名称
                    # print('直接进入 测试计划页面')
                    plan = '//*[text()=\"' + planName + '\"]'
                    print('测试计划为:%s' % planName)
                    time.sleep(2)
                    Test.getExistanceByxpath(xpath=plan)
                    # 判断是否有测试计划
                    if element_exist == True:
                        print('当前页测试计划存在，可正常点击！')
                    else:
                        print('当前页测试计划不存在，需翻页！')
                        # 配置每页数量
                        pageXpath = '//*[text()="每页显示:"]/../span[2]'
                        page = self.driver.find_element_by_xpath(pageXpath).click()
                        try:
                            pageMAXXpath = '/html/body/div[5]/div/ul/li[3]/div/span'
                            pageMAX = self.driver.find_element_by_xpath(pageMAXXpath).click()
                        except:
                            pageMAXXpath = '/html/body/div[4]/div/ul/li[3]/div/span'
                            pageMAX = self.driver.find_element_by_xpath(pageMAXXpath).click()
                        WebDriverWait(self.driver, 15).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[text()="计划名称"]')))  # 等待显示 计划名称
                        # 获取页数
                        pageNumXpath = '//*[@id="container"]/div/main/section/section/section/div/div/div/div/div[2]/div[2]/div[1]/div'
                        pageNumModel = self.driver.find_element_by_xpath(pageNumXpath)
                        pageNum = pageNumModel.find_elements_by_xpath('button')
                        for pNum in range(1, len(pageNum) + 1):
                            pageNumButtonXpath = pageNumXpath + f'/button[{pNum}]'
                            self.driver.find_element_by_xpath(pageNumButtonXpath).click()
                            WebDriverWait(self.driver, 15).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[text()="计划名称"]')))  # 等待显示 计划名称
                            Test.getExistanceByxpath(xpath=plan)
                            # print(element_exist)
                            # 判断是否有测试计划
                            if element_exist == True:
                                print(f'进入第{pNum}页，测试计划存在，可正常点击！')
                                break
                            else:
                                print(f'进入第{pNum}页未找到测试计划！')
                    # 选择计划
                    self.driver.find_element_by_xpath(plan).click()
                    # time.sleep(3)
                    Test.getCase(fileName=planName)
                    # 当前计划结束，结束计时
                    endTime = time.clock()
                    # 计算用时
                    print(f'测试导出引用的计划：{planName},用时：{int(endTime - startTime)}秒')
            # 开始编辑数据并引用用例
            filePath = f'./output/planCaseAuto/{autoPlanName} 测试结果记录表.xlsx'
            Test.editFile(fileNamePath=filePath)
            #结束计时
            endTime = time.clock()
            # 计算用时
            print(f'测试计划：{newPlanName},引用用例用时：{int(endTime - startTime)}秒')
            time.sleep(15)
            # 关闭浏览器
            self.driver.quit()
            print('结束关闭浏览器')


if __name__ == "__main__":
    import logging
    now_datetime = datetime.now()
    str_now_datetime = now_datetime.strftime("%Y-%m-%d %H-%M-%S")  # 获取时间
    print("now date & time : ", str_now_datetime)
    logName = './output/logs/DebugLog%s.txt' % str_now_datetime
    logging.basicConfig(filename=logName, level=logging.DEBUG)
    logging.debug('debug level test')
    logging.info('info level test')
    logging.warning('warning level test')
    logging.error('error level test')
    logging.critical('critical level test')


    Test = automaticLabeling()
    loginYX = Test.login()
    if loginYX == True:
        Test.action()
    # Test.selectCaseMode(fileName=autoPlanName)

    #     oldName = './output/logs/logPycharm.txt'
    #     now_datetime = datetime.now()
    #     str_now_datetime = now_datetime.strftime("%Y-%m-%d %H-%M-%S")  # 获取时间
    #     print("now date & time : ", str_now_datetime)
    #     newName = './output/logs/logPycharm%s.txt' % str_now_datetime
    #     shutil.copy(oldName, newName)  # oldName需要复制的文件 newName另存文件的名字





